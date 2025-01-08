from flask import (
    Blueprint, 
    render_template, 
    request, 
    jsonify, 
    current_app, 
    flash, 
    redirect, 
    url_for,
    session,
    send_file
)
from pymongo import MongoClient
from bson import ObjectId
import pandas as pd
from datetime import datetime
from werkzeug.utils import secure_filename
import os
import math
from collections import Counter
from io import BytesIO
from decimal import Decimal
import traceback
from functools import wraps  # Add this import
from datetime import datetime, timedelta  # Add timedelta import
import logging
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
from io import BytesIO
from openpyxl.styles import NamedStyle, Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import traceback
from dotenv import load_dotenv
# Create Blueprint
financials = Blueprint('financials', __name__, url_prefix='/financials')

# MongoDB connection
# Load environment variables
load_dotenv()

def get_db():
    """Get MongoDB database connection"""
    try:
        # Get MongoDB URI from environment variable
        mongo_uri = os.getenv('MONGO_URI')
        if not mongo_uri:
            raise ValueError("MongoDB URI not found in environment variables")
            
        # Create MongoDB client
        client = MongoClient(mongo_uri)
        
        # Return database instance
        return client.IAapplication
        
    except Exception as e:
        print(f"Database connection error: {str(e)}")
        raise


# Define login_required decorator before routes
def login_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if 'logged_in' not in session or 'user_id' not in session:
            flash('Please log in to access this feature', 'error')
            return redirect(url_for('login'))
        return f(*args, **kwargs)
    return decorated_function


def init_db():
    try:
        db = get_db()
        db.financial_clients.create_index([("user_id", 1)])
        db.trial_balances.create_index([("user_id", 1)])
        db.trial_balances.create_index([("client_id", 1), ("user_id", 1)])
        db.financial_clients.create_index([
        ("user_id", 1),
        ("pan", 1),
        ("fiscal_year", 1)
        ], unique=True)
        db.financial_clients.create_index([("created_at", -1)])
        db.financial_audit_log.create_index([("timestamp", -1)])
        db.financial_audit_log.create_index([("client_id", 1)])
    except Exception as e:
        print(f"Error initializing database: {str(e)}")
        raise



# Custom template filters

def get_bs_value(balance_sheet, path):
    """Safely get value from balance sheet using dot notation path."""
    try:
        if not balance_sheet:
            return 0
            
        value = balance_sheet
        for key in path.split('.'):
            value = value.get(key, {})
            
        if isinstance(value, dict):
            return value.get('total', 0)
        return float(value or 0)
    except Exception as e:
        print(f"Error getting balance sheet value for {path}: {str(e)}")
        return 0

def apply_number_style(worksheet, row, columns, bold=False):
    """
    Apply number formatting to specified cells.
    
    Args:
        worksheet: Excel worksheet object
        row (int): Row number
        columns (list): List of column letters
        bold (bool): Whether to apply bold formatting
    """
    for col in columns:
        cell = worksheet[f'{col}{row}']
        cell.number_format = '#,##0.00'
        if bold:
            cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='right')

# Excel styles
title_style = NamedStyle(
    name='title_style',
    font=Font(bold=True, size=14),
    alignment=Alignment(horizontal='center', vertical='center')
)

header_style = NamedStyle(
    name='header_style',
    font=Font(bold=True),
    fill=PatternFill(start_color='E0E0E0', end_color='E0E0E0', fill_type='solid'),
    alignment=Alignment(horizontal='left', vertical='center')
)

subheader_style = NamedStyle(
    name='subheader_style',
    font=Font(bold=True),
    alignment=Alignment(horizontal='left', vertical='center')
)

number_style = NamedStyle(
    name='number_style',
    number_format='#,##0.00',
    alignment=Alignment(horizontal='right', vertical='center')
)

@financials.app_template_filter()
def absolute(number):
    try:
        return abs(float(number))
    except (TypeError, ValueError):
        return 0.0

@financials.app_template_filter('datetime')
def format_datetime(value):
    if isinstance(value, datetime):
        return value.strftime('%Y-%m-%d %H:%M')
    return str(value)

# Now define your routes using the decorator
@financials.route('/')
@login_required
def financial_dashboard():
    try:
        db = get_db()
        # Fetch only clients belonging to the logged-in user
        clients = list(db.financial_clients.find(
            {'user_id': session['user_id']},
            # Specify the fields we want to retrieve
            {
                'company_name': 1,
                'legal_name': 1,
                'fiscal_year': 1,
                'status': 1,
                'updated_at': 1
            }
        ))

        # Transform the data to ensure company name is never None
        for client in clients:
            client['company_name'] = client.get('legal_name') or client.get('company_name') or 'Unnamed Client'

        return render_template('financials/dashboard.html', clients=clients)
    except Exception as e:
        flash(f'Error loading dashboard: {str(e)}', 'error')
        return redirect(url_for('login'))

import re
from datetime import datetime
from bson import ObjectId

@financials.route('/add_client', methods=['GET', 'POST'])
@login_required
def add_financial_client():
    if request.method == 'POST':
        try:
            db = get_db()
            
            # Validate required fields
            required_fields = [
                'legal_name', 'pan', 'registration_number', 'business_type',
                'registered_address', 'standards_applicability', 'signatory_name',
                'signatory_designation', 'signatory_pan', 'signatory_place',
                'signatory_date', 'audit_firm_name', 'frn', 'udin',
                'auditor_name', 'membership_no', 'caro_applicable',
                'rounding_amount', 'decimal_rounding', 'fiscal_year',
                'start_date', 'end_date'
            ]
            
            # Check for missing required fields
            missing_fields = []
            for field in required_fields:
                if not request.form.get(field):
                    missing_fields.append(field.replace('_', ' ').title())
            
            if missing_fields:
                flash(f'Missing required fields: {", ".join(missing_fields)}', 'error')
                return redirect(url_for('financials.add_financial_client'))

            # Validate PAN format
            pan_pattern = re.compile(r'^[A-Z]{5}[0-9]{4}[A-Z]{1}$')
            if not pan_pattern.match(request.form.get('pan')):
                flash('Invalid PAN format', 'error')
                return redirect(url_for('financials.add_financial_client'))

            if not pan_pattern.match(request.form.get('signatory_pan')):
                flash('Invalid Signatory PAN format', 'error')
                return redirect(url_for('financials.add_financial_client'))

            # Validate dates
            try:
                start_date = datetime.strptime(request.form.get('start_date'), '%Y-%m-%d')
                end_date = datetime.strptime(request.form.get('end_date'), '%Y-%m-%d')
                signatory_date = datetime.strptime(request.form.get('signatory_date'), '%Y-%m-%d')

                if start_date > end_date:
                    flash('Start date cannot be after end date', 'error')
                    return redirect(url_for('financials.add_financial_client'))

            except ValueError:
                flash('Invalid date format', 'error')
                return redirect(url_for('financials.add_financial_client'))

            # Validate business type
            valid_business_types = [
                'Private Limited', 'Individual', 'Partnership Firm',
                'Public Limited', 'Trust', 'One Person Company'
            ]
            if request.form.get('business_type') not in valid_business_types:
                flash('Invalid business type', 'error')
                return redirect(url_for('financials.add_financial_client'))

            # Validate standards applicability
            valid_standards = ['INDAS', 'IGAAP']
            if request.form.get('standards_applicability') not in valid_standards:
                flash('Invalid standards applicability', 'error')
                return redirect(url_for('financials.add_financial_client'))

            # Create client data document
            client_data = {
                # User and Status Information
                'user_id': session['user_id'],
                'status': 'Pending',
                'created_at': datetime.utcnow(),
                'updated_at': datetime.utcnow(),
                
                # Basic Information
                'legal_name': request.form.get('legal_name'),
                'pan': request.form.get('pan').upper(),
                'registration_number': request.form.get('registration_number'),
                'business_type': request.form.get('business_type'),
                'registered_address': request.form.get('registered_address'),
                
                # Financial Information
                'fiscal_year': request.form.get('fiscal_year'),
                'standards_applicability': request.form.get('standards_applicability'),
                'start_date': start_date,
                'end_date': end_date,
                
                # Signatory Details
                'signatory': {
                    'name': request.form.get('signatory_name'),
                    'designation': request.form.get('signatory_designation'),
                    'pan': request.form.get('signatory_pan').upper(),
                    'place': request.form.get('signatory_place'),
                    'date': signatory_date
                },
                
                # Audit Firm Details
                'audit_firm': {
                    'name': request.form.get('audit_firm_name'),
                    'frn': request.form.get('frn'),
                    'udin': request.form.get('udin'),
                    'auditor_name': request.form.get('auditor_name'),
                    'membership_no': request.form.get('membership_no'),
                    'place': request.form.get('signatory_place'),
                    'date': signatory_date
                },
                
                # Additional Settings
                'settings': {
                    'caro_applicable': request.form.get('caro_applicable') == 'Yes',
                    'rounding_amount': request.form.get('rounding_amount'),
                    'decimal_rounding': int(request.form.get('decimal_rounding'))
                },
                
                # Metadata
                'last_modified_by': session.get('username'),
                'version': 1,
                'is_active': True
            }

            # Check for duplicate client
            existing_client = db.financial_clients.find_one({
                'user_id': session['user_id'],
                'pan': client_data['pan'],
                'fiscal_year': client_data['fiscal_year']
            })

            if existing_client:
                flash('A client with this PAN and fiscal year already exists', 'error')
                return redirect(url_for('financials.add_financial_client'))

            # Insert into database
            result = db.financial_clients.insert_one(client_data)

            # Create audit log entry
            audit_log = {
                'user_id': session['user_id'],
                'client_id': result.inserted_id,
                'action': 'create_client',
                'timestamp': datetime.utcnow(),
                'details': {
                    'legal_name': client_data['legal_name'],
                    'fiscal_year': client_data['fiscal_year']
                },
                'ip_address': request.remote_addr,
                'user_agent': request.user_agent.string
            }
            db.financial_audit_log.insert_one(audit_log)

            flash(f'Client {client_data["legal_name"]} added successfully', 'success')
            return redirect(url_for('financials.financial_dashboard'))

        except Exception as e:
            # Log the error
            print(f"Error adding client: {str(e)}")
            flash(f'Error adding client: {str(e)}', 'error')
            return redirect(url_for('financials.add_financial_client'))

    # GET request - render the form
    return render_template('financials/add_client.html')

# Helper function to validate PAN
def is_valid_pan(pan):
    """Validate PAN number format."""
    if not pan:
        return False
    pan_pattern = re.compile(r'^[A-Z]{5}[0-9]{4}[A-Z]{1}$')
    return bool(pan_pattern.match(pan))

# Helper function to format currency values
def format_amount(amount, rounding_type, decimals):
    """Format amount based on rounding type and decimals."""
    try:
        divisor = {
            'Thousand': 1000,
            'Lakhs': 100000,
            'Million': 1000000,
            'Crores': 10000000
        }.get(rounding_type, 1)
        
        rounded_amount = round(float(amount) / divisor, decimals)
        
        # For negative numbers, show with minus sign
        if rounded_amount < 0:
            return f"({abs(rounded_amount):,.{decimals}f})"  # Show negative numbers in parentheses
        return f"{rounded_amount:,.{decimals}f}"
        
    except (TypeError, ValueError):
        return "0.00"

# Helper function to validate UDIN
def is_valid_udin(udin):
    """Validate UDIN format."""
    if not udin:
        return False
    # Add your UDIN validation logic here
    # Example: 20-digit alphanumeric
    udin_pattern = re.compile(r'^[A-Z0-9]{20}$')
    return bool(udin_pattern.match(udin))

# Helper function to validate FRN
def is_valid_frn(frn):
    """Validate Firm Registration Number format."""
    if not frn:
        return False
    # Add your FRN validation logic here
    # Example: Format like "123456W" or "123456E"
    frn_pattern = re.compile(r'^\d{6}[WE]$')
    return bool(frn_pattern.match(frn))

# Add the rest of your routes with @login_required decorator where needed
# ... (rest of your code remains the same)

# Define P&L mapping structure
# Define P&L mapping structure
PL_MAPPING_STRUCTURE = {
    'revenue_from_operations': {
        'subschedules': [
            'PL-sales_revenue',
            'PL-service_income'
        ],
        'section': 'income'
    },
    'other_income': {
        'subschedules': [
            'PL-interest_income',
            'PL-Other Income',
            'PL-other_operating_income'
        ],
        'section': 'income'
    },
    'cost_of_materials': {
        'subschedules': [
            'PL-material_purchases',
            'PL-Cost of Sales',
            'PL-direct_expenses'
        ],
        'section': 'expenses'
    },
    'employee_benefits': {
        'subschedules': [
            'PL-employee_benefits'
        ],
        'section': 'expenses'
    },
    'finance_costs': {
        'subschedules': [
            'PL-Financial Expenses',
            'PL-finance_costs'
        ],
        'section': 'expenses'
    },
    'depreciation_and_amortisation': {
        'subschedules': [
            'PL-depreciation'
        ],
        'section': 'expenses'
    },
    'other_expenses': {
        'subschedules': [
            'PL-Operating Expenses',
            'PL-admin_expenses',
            'PL-selling_expenses'
        ],
        'section': 'expenses'
    },
    'tax_expenses': {
        'subschedules': [
            'PL-Tax Expenses',
            'PL-current_tax',
            'PL-deferred_tax'
        ],
        'section': 'expenses'
    }
}

# Create reverse mapping for easy lookup
PL_SUBSCHEDULE_TO_MAIN = {}
for main_category, data in PL_MAPPING_STRUCTURE.items():
    for subschedule in data['subschedules']:
        PL_SUBSCHEDULE_TO_MAIN[subschedule] = {
            'main_category': main_category,
            'section': data['section']
        }

def get_main_pl_category(subschedule):
    """Helper function to get main P&L category from subschedule"""
    return PL_SUBSCHEDULE_TO_MAIN.get(subschedule, {}).get('main_category')

def get_pl_section(subschedule):
    """Helper function to get P&L section (income/expenses) from subschedule"""
    return PL_SUBSCHEDULE_TO_MAIN.get(subschedule, {}).get('section')

# Upload Trial Balance
@financials.route('/upload_trial_balance/<client_id>', methods=['GET', 'POST'])
@login_required
def upload_trial_balance(client_id):
    try:
        db = get_db()
        
        # Get client with user verification
        client = db.financial_clients.find_one({
            '_id': ObjectId(client_id),
            'user_id': session['user_id']
        })

        if not client:
            flash('Client not found or access denied', 'error')
            return redirect(url_for('financials.financial_dashboard'))

        # Handle GET request
        if request.method == 'GET':
            # Get existing trial balance if any
            trial_balance = db.trial_balances.find_one({
                'client_id': ObjectId(client_id),
                'user_id': session['user_id']
            })
            
            return render_template('financials/upload_trial_balance.html', 
                                 client=client,
                                 trial_balance=trial_balance)

        # Handle POST request
        if request.method == 'POST':
            # Check if file was uploaded
            if 'trial_balance_file' not in request.files:
                return jsonify({
                    'success': False,
                    'error': 'No file uploaded'
                }), 400
            
            file = request.files['trial_balance_file']
            
            # Check if a file was selected
            if file.filename == '':
                return jsonify({
                    'success': False,
                    'error': 'No file selected'
                }), 400
            
            # Check file extension
            if not file.filename.endswith(('.xlsx', '.xls')):
                return jsonify({
                    'success': False,
                    'error': 'Invalid file format. Please upload an Excel file (.xlsx or .xls)'
                }), 400

            # Read Excel file
            try:
                df = pd.read_excel(file)
            except Exception as e:
                return jsonify({
                    'success': False,
                    'error': f'Error reading Excel file: {str(e)}'
                }), 400

            # Validate required columns
            required_columns = ['Account Code', 'Account Name', 'Debit', 'Credit', 
                              'Account Type', 'Financial Statement Mapping']
            missing_columns = [col for col in required_columns if col not in df.columns]
            if missing_columns:
                return jsonify({
                    'success': False,
                    'error': f'Missing required columns: {", ".join(missing_columns)}'
                }), 400

            # Valid mapping options
            # Update in upload_trial_balance function
            valid_mappings = {
                # Current Assets subschedules
                'BS-cash_and_equivalents',
                'BS-trade_receivables',
                'BS-inventories',
                'BS-loans_and_advances',
                'BS-other_current_assets',
                
                # Fixed Assets subschedules
                'BS-tangible_assets',
                'BS-intangible_assets',
                
                # Investments subschedules
                'BS-long_term_investments',
                'BS-short_term_investments',
                
                # Liability subschedules
                'BS-msme_payables',
                'BS-other_payables',
                'BS-statutory_dues',
                'BS-short_term_borrowings',
                'BS-other_current_liabilities',
                'BS-secured_loans',
                'BS-unsecured_loans',
                
                # Equity subschedules
                'BS-equity_share_capital',
                'BS-general_reserve',
                'BS-capital_reserve',
                'BS-retained_earnings',
                'BS-securities_premium',
                
                # P&L subschedules - Income
                'PL-sales_revenue',
                'PL-service_income',
                'PL-interest_income',
                'PL-Other Income',
                'PL-other_operating_income',
                
                # P&L subschedules - Expenses
                'PL-material_purchases',
                'PL-Cost of Sales',
                'PL-Operating Expenses',
                'PL-Financial Expenses',
                'PL-Tax Expenses',
                'PL-direct_expenses',
                'PL-employee_benefits',
                'PL-finance_costs',
                'PL-depreciation',
                'PL-admin_expenses',
                'PL-selling_expenses',
                'PL-current_tax',
                'PL-deferred_tax'
            }

            # Valid account types
            valid_account_types = {'Asset', 'Liability', 'Equity', 'Revenue', 'Expense'}

            # Clean and process the data
            trial_balance_data = []
            total_debit = 0
            total_credit = 0

            for index, row in df.iterrows():
                # Validate Account Code and Name
                if pd.isna(row['Account Code']) or pd.isna(row['Account Name']):
                    return jsonify({
                        'success': False,
                        'error': f'Missing Account Code or Name at row {index + 2}'
                    }), 400

                # Clean and validate Account Code
                account_code = str(row['Account Code']).strip()
                if not account_code:
                    return jsonify({
                        'success': False,
                        'error': f'Invalid Account Code at row {index + 2}'
                    }), 400

                # Convert and validate numeric values
                try:
                    debit = float(row['Debit']) if pd.notna(row['Debit']) else 0.0
                    credit = float(row['Credit']) if pd.notna(row['Credit']) else 0.0
                    
                    # Validate non-negative values
                    if debit < 0 or credit < 0:
                        return jsonify({
                            'success': False,
                            'error': f'Negative values not allowed in Debit/Credit at row {index + 2}'
                        }), 400
                    
                    total_debit += debit
                    total_credit += credit
                except ValueError:
                    return jsonify({
                        'success': False,
                        'error': f'Invalid numeric value in Debit or Credit at row {index + 2}'
                    }), 400

                # Validate Account Type
                account_type = str(row['Account Type']).strip()
                if account_type not in valid_account_types:
                    return jsonify({
                        'success': False,
                        'error': f'Invalid Account Type at row {index + 2}. Must be one of: {", ".join(valid_account_types)}'
                    }), 400

                # Validate Financial Statement Mapping
                mapping = str(row['Financial Statement Mapping']).strip()
                if mapping not in valid_mappings:
                    return jsonify({
                        'success': False,
                        'error': f'Invalid Financial Statement Mapping at row {index + 2}. Must be one of: {", ".join(valid_mappings)}'
                    }), 400

                trial_balance_data.append({
                    'Account_Code': account_code,
                    'Account_Name': str(row['Account Name']).strip(),
                    'Debit': debit,
                    'Credit': credit,
                    'Account_Type': account_type,
                    'Mapping': mapping
                })

            # Validate trial balance
            if not math.isclose(total_debit, total_credit, rel_tol=1e-9):
                return jsonify({
                    'success': False,
                    'error': f'Trial balance is not balanced. Total Debit: {total_debit:,.2f}, Total Credit: {total_credit:,.2f}'
                }), 400

            # Check for duplicate account codes
            account_codes = [item['Account_Code'] for item in trial_balance_data]
            duplicates = [code for code, count in Counter(account_codes).items() if count > 1]
            if duplicates:
                return jsonify({
                    'success': False,
                    'error': f'Duplicate account codes found: {", ".join(duplicates)}'
                }), 400

            try:
                # Prepare trial balance document
                trial_balance_doc = {
                    'client_id': ObjectId(client_id),
                    'user_id': session['user_id'],
                    'data': trial_balance_data,
                    'total_debit': total_debit,
                    'total_credit': total_credit,
                    'uploaded_at': datetime.utcnow(),
                    'file_name': file.filename,
                    'status': 'Active',
                    'period': {
                        'start_date': client['start_date'],
                        'end_date': client['end_date']
                    }
                }

                # Update or insert trial balance
                db.trial_balances.update_one(
                    {
                        'client_id': ObjectId(client_id),
                        'user_id': session['user_id']
                    },
                    {'$set': trial_balance_doc},
                    upsert=True
                )

                # Update client status
                db.financial_clients.update_one(
                    {
                        '_id': ObjectId(client_id),
                        'user_id': session['user_id']
                    },
                    {
                        '$set': {
                            'status': 'Uploaded',
                            'updated_at': datetime.utcnow(),
                            'last_upload_date': datetime.utcnow()
                        }
                    }
                )

                # Create audit log entry
                audit_log = {
                    'client_id': ObjectId(client_id),
                    'user_id': session['user_id'],
                    'action': 'upload_trial_balance',
                    'details': {
                        'file_name': file.filename,
                        'total_entries': len(trial_balance_data),
                        'total_debit': total_debit,
                        'total_credit': total_credit
                    },
                    'timestamp': datetime.utcnow(),
                    'ip_address': request.remote_addr,
                    'user_agent': request.user_agent.string
                }
                db.financial_audit_log.insert_one(audit_log)

                return jsonify({
                    'success': True,
                    'message': 'Trial balance uploaded successfully',
                    'details': {
                        'total_entries': len(trial_balance_data),
                        'total_debit': total_debit,
                        'total_credit': total_credit
                    }
                })

            except Exception as e:
                print(f"Database error: {str(e)}")
                return jsonify({
                    'success': False,
                    'error': 'Error saving trial balance to database'
                }), 500

    except Exception as e:
        print(f"Error processing trial balance: {str(e)}")
        return jsonify({
            'success': False,
            'error': 'An error occurred while processing the trial balance'
        }), 500

    return jsonify({
        'success': False,
        'error': 'Invalid request method'
    }), 405

def allowed_file(filename):
    ALLOWED_EXTENSIONS = {'xlsx', 'xls'}
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

def process_trial_balance(df):
    # Process the trial balance DataFrame
    # Validate required columns
    required_columns = ['Account Code', 'Account Name', 'Debit', 'Credit']
    if not all(col in df.columns for col in required_columns):
        raise ValueError("Missing required columns in trial balance")
    
    # Convert to dictionary format
    trial_balance_data = df.to_dict('records')
    
    # Validate debits and credits balance
    total_debit = df['Debit'].sum()
    total_credit = df['Credit'].sum()
    if not abs(total_debit - total_credit) < 0.01:  # Allow small rounding differences
        raise ValueError("Trial balance is not balanced")
    
    return trial_balance_data

@financials.route('/download_template')
def download_template():
    try:
        # Create DataFrame with sample mappings
        template_data = {
            'Account Code': [],
            'Account Name': [],
            'Debit': [],
            'Credit': [],
            'Account Type': [],
            'Financial Statement Mapping': []
        }
        
        # Define comprehensive sample accounts with sub-schedule mappings
        sample_accounts = [
            # Current Assets sub-schedules
            ('1001', 'Cash in Hand', 'Asset', 'BS-cash_and_equivalents', 50000, 0),
            ('1002', 'Bank Balances', 'Asset', 'BS-cash_and_equivalents', 150000, 0),
            ('1003', 'Trade Receivables', 'Asset', 'BS-trade_receivables', 75000, 0),
            ('1004', 'Short Term Loans', 'Asset', 'BS-loans_and_advances', 25000, 0),
            ('1005', 'Inventory', 'Asset', 'BS-inventories', 100000, 0),
            
            # Fixed Assets sub-schedules
            ('2001', 'Plant and Machinery', 'Asset', 'BS-tangible_assets', 1130000, 0),
            ('2002', 'Office Equipment', 'Asset', 'BS-tangible_assets', 100000, 0),
            ('2003', 'Computer Software', 'Asset', 'BS-intangible_assets', 50000, 0),
            
            # Investments sub-schedules
            ('3001', 'Long Term Investments', 'Asset', 'BS-long_term_investments', 200000, 0),
            ('3002', 'Short Term Investments', 'Asset', 'BS-short_term_investments', 100000, 0),
            
            # Current Liabilities sub-schedules
            ('4001', 'Trade Payables - MSME', 'Liability', 'BS-msme_payables', 0, 50000),
            ('4002', 'Trade Payables - Others', 'Liability', 'BS-other_payables', 0, 150000),
            ('4003', 'GST Payable', 'Liability', 'BS-statutory_dues', 0, 25000),
            ('4004', 'Short Term Borrowings', 'Liability', 'BS-short_term_borrowings', 0, 100000),
            
            # Long Term Liabilities sub-schedules
            ('5001', 'Term Loan', 'Liability', 'BS-secured_loans', 0, 300000),
            ('5002', 'Unsecured Loans', 'Liability', 'BS-unsecured_loans', 0, 200000),
            
            # Equity sub-schedules
            ('6001', 'Equity Share Capital', 'Equity', 'BS-equity_share_capital', 0, 500000),
            ('6002', 'General Reserve', 'Equity', 'BS-general_reserve', 0, 100000),
            ('6003', 'Capital Reserve', 'Equity', 'BS-capital_reserve', 0, 50000),
            ('6004', 'Retained Earnings', 'Equity', 'BS-retained_earnings', 0, 75000),

            # Revenue accounts
            ('7001', 'Sales Revenue', 'Revenue', 'PL-sales_revenue', 0, 1150000),
            ('7002', 'Service Income', 'Revenue', 'PL-service_income', 0, 200000),
            ('7003', 'Interest Income', 'Revenue', 'PL-interest_income', 0, 50000),
            ('7004', 'Other Operating Income', 'Revenue', 'PL-other_operating_income', 0, 30000),
            
            # Cost of Materials
            ('7101', 'Raw Material Purchases', 'Expense', 'PL-material_purchases', 300000, 0),
            ('7102', 'Direct Expenses', 'Expense', 'PL-direct_expenses', 50000, 0),
            ('7103', 'Production Costs', 'Expense', 'PL-Cost of Sales', 150000, 0),
            
            # Employee Benefits
            ('7201', 'Salaries and Wages', 'Expense', 'PL-employee_benefits', 200000, 0),
            ('7202', 'Staff Welfare', 'Expense', 'PL-employee_benefits', 30000, 0),
            
            # Finance Costs
            ('7301', 'Bank Interest', 'Expense', 'PL-finance_costs', 25000, 0),
            ('7302', 'Loan Processing Charges', 'Expense', 'PL-Financial Expenses', 5000, 0),
            
            # Depreciation
            ('7401', 'Depreciation on Fixed Assets', 'Expense', 'PL-depreciation', 75000, 0),
            
            # Other Expenses
            ('7501', 'Administrative Expenses', 'Expense', 'PL-admin_expenses', 40000, 0),
            ('7502', 'Selling Expenses', 'Expense', 'PL-selling_expenses', 35000, 0),
            ('7503', 'General Operating Expenses', 'Expense', 'PL-Operating Expenses', 45000, 0),
            
            # Tax Expenses
            ('7601', 'Current Tax', 'Expense', 'PL-current_tax', 40000, 0),
            ('7602', 'Deferred Tax', 'Expense', 'PL-deferred_tax', 5000, 0)
        ]
        # Add accounts to template
        for code, name, type_, mapping, debit, credit in sample_accounts:
            template_data['Account Code'].append(code)
            template_data['Account Name'].append(name)
            template_data['Account Type'].append(type_)
            template_data['Financial Statement Mapping'].append(mapping)
            template_data['Debit'].append(debit)
            template_data['Credit'].append(credit)

        template_df = pd.DataFrame(template_data)
        
        # Create Excel file with formatting
        output = BytesIO()
        with pd.ExcelWriter(output, engine='xlsxwriter') as writer:
            template_df.to_excel(writer, index=False, sheet_name='Trial Balance')
            workbook = writer.book
            worksheet = writer.sheets['Trial Balance']

            # Define formats
            header_format = workbook.add_format({
                'bold': True,
                'bg_color': '#D3D3D3',
                'border': 1,
                'text_wrap': True,
                'align': 'center',
                'valign': 'vcenter'
            })
            
            number_format = workbook.add_format({
                'num_format': '#,##0.00',
                'border': 1,
                'align': 'right'
            })
            
            text_format = workbook.add_format({
                'border': 1,
                'align': 'left'
            })

            # Format headers
            for col_num, value in enumerate(template_df.columns.values):
                worksheet.write(0, col_num, value, header_format)
                
            # Format data
            for row_num in range(len(template_df)):
                worksheet.write(row_num + 1, 0, template_df.iloc[row_num, 0], text_format)  # Account Code
                worksheet.write(row_num + 1, 1, template_df.iloc[row_num, 1], text_format)  # Account Name
                worksheet.write(row_num + 1, 2, template_df.iloc[row_num, 2], number_format)  # Debit
                worksheet.write(row_num + 1, 3, template_df.iloc[row_num, 3], number_format)  # Credit
                worksheet.write(row_num + 1, 4, template_df.iloc[row_num, 4], text_format)  # Account Type
                worksheet.write(row_num + 1, 5, template_df.iloc[row_num, 5], text_format)  # Mapping

            # Set column widths
            worksheet.set_column('A:A', 15)  # Account Code
            worksheet.set_column('B:B', 30)  # Account Name
            worksheet.set_column('C:D', 15)  # Debit/Credit
            worksheet.set_column('E:E', 15)  # Account Type
            worksheet.set_column('F:F', 25)  # Financial Statement Mapping

            # Add validation for Account Type
            account_types = ['Asset', 'Liability', 'Equity', 'Revenue', 'Expense']
            worksheet.data_validation('E2:E1048576', {
                'validate': 'list',
                'source': account_types,
                'input_title': 'Select Account Type',
                'input_message': 'Please select from the list of valid account types'
            })

            # Create hidden validation sheet for mappings
            validation_sheet = workbook.add_worksheet('_ValidationLists')
            validation_sheet.hide()

            # Valid mapping options
            valid_mappings = [
                # Balance Sheet Mappings
                'BS-cash_and_equivalents',
                'BS-trade_receivables',
                'BS-inventories',
                'BS-loans_and_advances',
                'BS-other_current_assets',
                'BS-tangible_assets',
                'BS-intangible_assets',
                'BS-long_term_investments',
                'BS-short_term_investments',
                'BS-msme_payables',
                'BS-other_payables',
                'BS-statutory_dues',
                'BS-short_term_borrowings',
                'BS-other_current_liabilities',
                'BS-secured_loans',
                'BS-unsecured_loans',
                'BS-equity_share_capital',
                'BS-general_reserve',
                'BS-capital_reserve',
                'BS-retained_earnings',
                # P&L Mappings
                'PL-sales_revenue',
                'PL-service_income',
                'PL-Other Income',
                'PL-Cost of Sales',
                'PL-Operating Expenses',
                'PL-Financial Expenses',
                'PL-Tax Expenses'
            ]

            # Write mappings to validation sheet
            for i, mapping in enumerate(valid_mappings):
                validation_sheet.write(i, 0, mapping)

            # Add validation to Mapping column
            worksheet.data_validation('F2:F1048576', {
                'validate': 'list',
                'source': '=_ValidationLists!$A$1:$A$' + str(len(valid_mappings)),
                'input_title': 'Select Mapping',
                'input_message': 'Please select from the list of valid mappings'
            })

            # Add Guide sheet
            guide_sheet = workbook.add_worksheet('Mapping Guide')
            guide_format = workbook.add_format({
                'bold': True,
                'bg_color': '#E6E6E6',
                'border': 1
            })
            
            # Write Balance Sheet mappings guide
            guide_sheet.write(0, 0, 'Balance Sheet Mappings', guide_format)
            bs_mappings = [
                ('BS-cash_and_equivalents', 'Cash, bank balances and equivalents'),
                ('BS-trade_receivables', 'Trade debtors and receivables'),
                ('BS-inventories', 'Stock and inventory items'),
                ('BS-loans_and_advances', 'Short-term loans and advances'),
                ('BS-other_current_assets', 'Other current assets'),
                ('BS-tangible_assets', 'Plant, property and equipment'),
                ('BS-intangible_assets', 'Software and other intangible assets'),
                ('BS-long_term_investments', 'Long term investments'),
                ('BS-short_term_investments', 'Short term investments'),
                ('BS-msme_payables', 'Trade payables to MSME vendors'),
                ('BS-other_payables', 'Other trade payables'),
                ('BS-statutory_dues', 'Statutory liabilities and dues'),
                ('BS-short_term_borrowings', 'Short term loans and borrowings'),
                ('BS-secured_loans', 'Long term secured loans'),
                ('BS-unsecured_loans', 'Long term unsecured loans'),
                ('BS-equity_share_capital', 'Equity share capital'),
                ('BS-general_reserve', 'General reserves'),
                ('BS-capital_reserve', 'Capital reserves'),
                ('BS-retained_earnings', 'Accumulated profits/losses')
            ]
            
            row = 1
            for mapping, desc in bs_mappings:
                guide_sheet.write(row, 0, mapping)
                guide_sheet.write(row, 1, desc)
                row += 1

            # Write P&L mappings guide
            guide_sheet.write(row + 1, 0, 'Profit & Loss Mappings', guide_format)
            pl_mappings = [
                ('PL-Revenue', 'Operating revenue and sales'),
                ('PL-Other Income', 'Non-operating income'),
                ('PL-Cost of Sales', 'Direct costs and purchases'),
                ('PL-Operating Expenses', 'Operating and administrative expenses'),
                ('PL-Financial Expenses', 'Interest and finance charges'),
                ('PL-Tax Expenses', 'Income tax and other taxes')
            ]
            
            row += 2
            for mapping, desc in pl_mappings:
                guide_sheet.write(row, 0, mapping)
                guide_sheet.write(row, 1, desc)
                row += 1

            # Set guide sheet column widths
            guide_sheet.set_column('A:A', 25)
            guide_sheet.set_column('B:B', 50)

        output.seek(0)
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name='trial_balance_template.xlsx'
        )

    except Exception as e:
        print(f"Template generation error: {str(e)}")
        flash('Error generating template file', 'error')
        return redirect(url_for('financials.financial_dashboard'))
        
@financials.route('/download_sample')
def download_sample():
    import pandas as pd
    from io import BytesIO
    
    # Sample data with mappings
    sample_data = {
        'Account Code': ['1001', '1002', '2001', '2002', '3001', '3002'],
        'Account Name': [
            'Cash and Bank', 'Trade Receivables', 
            'Current Liabilities', 'Long Term Loans',
            'Sales Revenue', 'Operating Expenses'
        ],
        'Debit': [100000, 50000, 0, 0, 0, 30000],
        'Credit': [0, 0, 40000, 60000, 80000, 0],
        'Account Type': [
            'Asset', 'Asset', 'Liability', 'Liability',
            'Revenue', 'Expense'
        ],
        'Financial Statement Mapping': [
            'BS-Current Assets', 'BS-Current Assets',
            'BS-Current Liabilities', 'BS-Long Term Liabilities',
            'PL-Revenue', 'PL-Operating Expenses'
        ]
    }
    
    df = pd.DataFrame(sample_data)
    
    # Rest of the code remains the same...
    
    # Create Excel file
    output = BytesIO()
    with pd.ExcelWriter(output, engine='xlsxwriter') as writer:
        df.to_excel(writer, index=False, sheet_name='Trial Balance')
        workbook = writer.book
        worksheet = writer.sheets['Trial Balance']
        
        # Add formatting
        header_format = workbook.add_format({
            'bold': True,
            'bg_color': '#D3D3D3',
            'border': 1
        })
        
        number_format = workbook.add_format({
            'num_format': '#,##0.00',
            'border': 1
        })
        
        text_format = workbook.add_format({
            'border': 1
        })
        
        # Format headers and data
        for col_num, value in enumerate(df.columns.values):
            worksheet.write(0, col_num, value, header_format)
            
        # Format data
        for row_num in range(len(df)):
            worksheet.write(row_num + 1, 0, df.iloc[row_num, 0], text_format)  # Account Code
            worksheet.write(row_num + 1, 1, df.iloc[row_num, 1], text_format)  # Account Name
            worksheet.write(row_num + 1, 2, df.iloc[row_num, 2], number_format)  # Debit
            worksheet.write(row_num + 1, 3, df.iloc[row_num, 3], number_format)  # Credit
            worksheet.write(row_num + 1, 4, df.iloc[row_num, 4], text_format)  # Account Type
            worksheet.write(row_num + 1, 5, df.iloc[row_num, 5], text_format)  # Mapping
            
        # Set column widths
        worksheet.set_column('A:A', 15)  # Account Code
        worksheet.set_column('B:B', 30)  # Account Name
        worksheet.set_column('C:D', 15)  # Debit/Credit
        worksheet.set_column('E:E', 15)  # Account Type
        worksheet.set_column('F:F', 25)  # Financial Statement Mapping
        
        # Add totals
        total_row = len(df) + 1
        worksheet.write(total_row, 1, 'Totals', header_format)
        worksheet.write_formula(total_row, 2, f'=SUM(C2:C{total_row})', number_format)
        worksheet.write_formula(total_row, 3, f'=SUM(D2:D{total_row})', number_format)
        
    output.seek(0)
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name='trial_balance_sample.xlsx'
    )
# In financials.py

@financials.route('/client/<client_id>/trial_balance')
def view_trial_balance(client_id):
    db = get_db()
    
    # Get client and trial balance data
    client = db.financial_clients.find_one({'_id': ObjectId(client_id)})
    trial_balance = db.trial_balances.find_one({'client_id': ObjectId(client_id)})
    
    if not client:
        flash('Client not found', 'error')
        return redirect(url_for('financials.financial_dashboard'))
        
    return render_template('financials/view_trial_balance.html', 
                         client=client, 
                         trial_balance=trial_balance)
@financials.route('/client/<client_id>/edit', methods=['GET', 'POST'])
@login_required
def edit_client(client_id):
    try:
        db = get_db()
        client = db.financial_clients.find_one({
            '_id': ObjectId(client_id),
            'user_id': session['user_id']
        })

        if not client:
            flash('Client not found', 'error')
            return redirect(url_for('financials.financial_dashboard'))

        if request.method == 'POST':
            # Convert dates from string to datetime
            try:
                start_date = datetime.strptime(request.form.get('start_date'), '%Y-%m-%d')
                end_date = datetime.strptime(request.form.get('end_date'), '%Y-%m-%d')
            except ValueError:
                flash('Invalid date format', 'error')
                return redirect(url_for('financials.edit_client', client_id=client_id))

            updated_data = {
                'legal_name': request.form.get('legal_name'),
                'pan': request.form.get('pan').upper(),
                'registration_number': request.form.get('registration_number'),
                'business_type': request.form.get('business_type'),
                'registered_address': request.form.get('registered_address'),
                'fiscal_year': request.form.get('fiscal_year'),
                'standards_applicability': request.form.get('standards_applicability'),
                'start_date': start_date,  # Add this
                'end_date': end_date,      # Add this
                'signatory': {
                    'name': request.form.get('signatory_name'),
                    'designation': request.form.get('signatory_designation'),
                    'pan': request.form.get('signatory_pan').upper(),
                    'place': request.form.get('signatory_place'),
                    'date': datetime.strptime(request.form.get('signatory_date'), '%Y-%m-%d') if request.form.get('signatory_date') else None
                },
                'audit_firm': {
                    'name': request.form.get('audit_firm_name'),
                    'frn': request.form.get('frn'),
                    'udin': request.form.get('udin'),
                    'auditor_name': request.form.get('auditor_name'),
                    'membership_no': request.form.get('membership_no')
                },
                'settings': {
                    'caro_applicable': request.form.get('caro_applicable') == 'Yes',
                    'rounding_amount': request.form.get('rounding_amount'),
                    'decimal_rounding': int(request.form.get('decimal_rounding'))
                },
                'updated_at': datetime.utcnow(),
                'last_modified_by': session.get('username')
            }

            result = db.financial_clients.update_one(
                {'_id': ObjectId(client_id)},
                {'$set': updated_data}
            )

            if result.modified_count > 0:
                flash('Client updated successfully', 'success')
            else:
                flash('No changes made', 'info')
            
            return redirect(url_for('financials.view_client_details', client_id=client_id))

        # For GET request, format dates for display
        if isinstance(client.get('start_date'), datetime):
            client['start_date'] = client['start_date'].strftime('%Y-%m-%d')
        if isinstance(client.get('end_date'), datetime):
            client['end_date'] = client['end_date'].strftime('%Y-%m-%d')

        return render_template('financials/edit_client.html', client=client)

    except Exception as e:
        print(f"Error updating client: {str(e)}")
        flash(f'Error updating client: {str(e)}', 'error')
        return redirect(url_for('financials.financial_dashboard'))

@financials.route('/client/<client_id>', methods=['DELETE'])
def delete_client(client_id):
    if 'logged_in' not in session or 'user_id' not in session:
        return jsonify({'success': False, 'error': 'Authentication required'}), 401

    try:
        db = get_db()
        # Add user_id check to ensure client belongs to logged-in user
        result = db.financial_clients.delete_one({
            '_id': ObjectId(client_id),
            'user_id': session['user_id']
        })

        if result.deleted_count:
            # Delete associated trial balance
            db.trial_balances.delete_one({
                'client_id': ObjectId(client_id),
                'user_id': session['user_id']
            })
            return jsonify({'success': True})
        else:
            return jsonify({'success': False, 'error': 'Client not found or access denied'}), 404
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

# In financials.py

from datetime import datetime
from bson import ObjectId

@financials.route('/client/<client_id>/details')
def view_client_details(client_id):
    if 'logged_in' not in session or 'user_id' not in session:
        flash('Please log in to view client details', 'error')
        return redirect(url_for('login'))

    try:
        db = get_db()
        # Add user_id check to ensure client belongs to logged-in user
        client = db.financial_clients.find_one({
            '_id': ObjectId(client_id),
            'user_id': session['user_id']
        })
        
        if not client:
            flash('Client not found or access denied', 'error')
            return redirect(url_for('financials.financial_dashboard'))

        # Convert string dates to datetime objects if they're strings
        if client.get('start_date') and isinstance(client['start_date'], str):
            client['start_date'] = datetime.strptime(client['start_date'], '%Y-%m-%d')
        if client.get('end_date') and isinstance(client['end_date'], str):
            client['end_date'] = datetime.strptime(client['end_date'], '%Y-%m-%d')

        # Get trial balance data if it exists
        trial_balance = db.trial_balances.find_one({
            'client_id': ObjectId(client_id),
            'user_id': session['user_id']  # Add user_id check for trial balance
        })

        return render_template('financials/client_details.html', 
                             client=client, 
                             trial_balance=trial_balance)
                             
    except Exception as e:
        print(f"Error in view_client_details: {str(e)}")
        flash(f'Error accessing client details: {str(e)}', 'error')
        return redirect(url_for('financials.financial_dashboard'))



from decimal import Decimal
from flask import flash, redirect, url_for, render_template
from bson import ObjectId

@financials.route('/generate_financial_statements/<client_id>')
@login_required
def generate_financial_statements(client_id):
    """Redirect to the new balance sheet generation"""
    return redirect(url_for('financials.generate_balance_sheet_view', client_id=client_id))

@financials.route('/export_financial_statements/<client_id>')
def export_financial_statements(client_id):
    try:
        db = get_db()
        client = db.financial_clients.find_one({'_id': ObjectId(client_id)})
        trial_balance = db.trial_balances.find_one({'client_id': ObjectId(client_id)})

        if not trial_balance or not client:
            flash('Data not found', 'error')
            return redirect(url_for('financials.financial_dashboard'))

        # Generate schedules (copy this from generate_financial_statements function)
        schedules = {
            'BS': {
                'Fixed Assets': [],
                'Current Assets': [],
                'Investments': [],
                'Other Assets': [],
                'Current Liabilities': [],
                'Long Term Liabilities': [],
                'Other Liabilities': [],
                'Share Capital': [],
                'Reserves': [],
                'Retained Earnings': []
            },
            'PL': {
                'Revenue': [],
                'Other Income': [],
                'Cost of Sales': [],
                'Operating Expenses': [],
                'Financial Expenses': [],
                'Tax Expenses': []
            }
        }

        # Process trial balance entries
        for entry in trial_balance.get('data', []):
            mapping = entry.get('Mapping', '')
            debit = float(entry.get('Debit', 0))
            credit = float(entry.get('Credit', 0))
            amount = debit - credit
            abs_amount = float(amount) if amount > 0 else float(-amount)

            if mapping.startswith('BS-'):
                category = mapping[3:]  # Remove 'BS-' prefix
                if category in schedules['BS']:
                    schedules['BS'][category].append({
                        'account_code': entry.get('Account_Code', ''),
                        'account_name': entry.get('Account_Name', ''),
                        'amount': abs_amount,
                        'type': 'Dr' if amount > 0 else 'Cr'
                    })
            
            elif mapping.startswith('PL-'):
                category = mapping[3:]  # Remove 'PL-' prefix
                if category in schedules['PL']:
                    schedules['PL'][category].append({
                        'account_code': entry.get('Account_Code', ''),
                        'account_name': entry.get('Account_Name', ''),
                        'amount': abs_amount,
                        'type': 'Dr' if amount > 0 else 'Cr'
                    })

        output = BytesIO()
        with pd.ExcelWriter(output, engine='xlsxwriter') as writer:
            workbook = writer.book
            
            # Define formats
            header_format = workbook.add_format({
                'bold': True,
                'bg_color': '#D3D3D3',
                'border': 1,
                'align': 'center'
            })
            
            title_format = workbook.add_format({
                'bold': True,
                'font_size': 14,
                'align': 'center'
            })
            
            normal_format = workbook.add_format({
                'border': 1
            })
            
            number_format = workbook.add_format({
                'border': 1,
                'num_format': '#,##0.00'
            })
            
            total_format = workbook.add_format({
                'bold': True,
                'border': 1,
                'bg_color': '#F0F0F0',
                'num_format': '#,##0.00'
            })

            # Create Balance Sheet
            bs_sheet = workbook.add_worksheet('Balance Sheet')
            bs_sheet.merge_range('A1:E1', f'{client["company_name"]} - Balance Sheet', title_format)
            bs_sheet.merge_range('A2:E2', f'As at {datetime.now().strftime("%d-%m-%Y")}', title_format)
            
            # Write headers
            headers = ['Account Code', 'Account Name', 'Description', 'Dr', 'Cr']
            for col, header in enumerate(headers):
                bs_sheet.write(3, col, header, header_format)
            
            current_row = 4
            for category, entries in schedules['BS'].items():
                if entries:  # Only write categories that have entries
                    bs_sheet.merge_range(
                        current_row, 0, current_row, 4,
                        category, 
                        workbook.add_format({'bold': True, 'bg_color': '#E6E6E6'})
                    )
                    current_row += 1
                    
                    for entry in entries:
                        bs_sheet.write(current_row, 0, entry['account_code'], normal_format)
                        bs_sheet.write(current_row, 1, entry['account_name'], normal_format)
                        amount = entry['amount']
                        if entry['type'] == 'Dr':
                            bs_sheet.write(current_row, 3, amount, number_format)
                            bs_sheet.write(current_row, 4, 0, number_format)
                        else:
                            bs_sheet.write(current_row, 3, 0, number_format)
                            bs_sheet.write(current_row, 4, amount, number_format)
                        current_row += 1
                    
                    # Category total
                    bs_sheet.write(
                        current_row, 2, 
                        f'Total {category}',
                        total_format
                    )
                    bs_sheet.write_formula(
                        current_row, 3,
                        f'=SUM(D{current_row-len(entries)}:D{current_row})',
                        total_format
                    )
                    bs_sheet.write_formula(
                        current_row, 4,
                        f'=SUM(E{current_row-len(entries)}:E{current_row})',
                        total_format
                    )
                    current_row += 2

            # Create P&L Sheet
            pl_sheet = workbook.add_worksheet('Profit and Loss')
            pl_sheet.merge_range('A1:E1', f'{client["company_name"]} - Profit & Loss Statement', title_format)
            pl_sheet.merge_range('A2:E2', f'For the period ending {datetime.now().strftime("%d-%m-%Y")}', title_format)
            
            # Write headers
            for col, header in enumerate(headers):
                pl_sheet.write(3, col, header, header_format)

            current_row = 4
            for category, entries in schedules['PL'].items():
                if entries:
                    pl_sheet.merge_range(
                        current_row, 0, current_row, 4,
                        category,
                        workbook.add_format({'bold': True, 'bg_color': '#E6E6E6'})
                    )
                    current_row += 1
                    
                    for entry in entries:
                        pl_sheet.write(current_row, 0, entry['account_code'], normal_format)
                        pl_sheet.write(current_row, 1, entry['account_name'], normal_format)
                        amount = entry['amount']
                        if entry['type'] == 'Dr':
                            pl_sheet.write(current_row, 3, amount, number_format)
                            pl_sheet.write(current_row, 4, 0, number_format)
                        else:
                            pl_sheet.write(current_row, 3, 0, number_format)
                            pl_sheet.write(current_row, 4, amount, number_format)
                        current_row += 1
                    
                    # Category total
                    pl_sheet.write(
                        current_row, 2,
                        f'Total {category}',
                        total_format
                    )
                    pl_sheet.write_formula(
                        current_row, 3,
                        f'=SUM(D{current_row-len(entries)}:D{current_row})',
                        total_format
                    )
                    pl_sheet.write_formula(
                        current_row, 4,
                        f'=SUM(E{current_row-len(entries)}:E{current_row})',
                        total_format
                    )
                    current_row += 2

            # Set column widths
            for sheet in [bs_sheet, pl_sheet]:
                sheet.set_column('A:A', 15)  # Account Code
                sheet.set_column('B:B', 30)  # Account Name
                sheet.set_column('C:C', 20)  # Description
                sheet.set_column('D:E', 15)  # Dr/Cr amounts

        output.seek(0)
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=f'{client["company_name"]}_Financial_Statements.xlsx'
        )

    except Exception as e:
        print(f"Export error: {str(e)}")  # For debugging
        flash(f'Error exporting financial statements: {str(e)}', 'error')
        return redirect(url_for('financials.financial_dashboard'))

def get_empty_balance_sheet():
    """Return empty balance sheet structure with all categories"""
    return {
        'assets': {
            'current_assets': {
                'cash_and_equivalents': {'entries': [], 'total': 0},
                'trade_receivables': {'entries': [], 'total': 0},
                'inventories': {'entries': [], 'total': 0},
                'loans_and_advances': {'entries': [], 'total': 0},
                'other_current_assets': {'entries': [], 'total': 0},
                'total': 0
            },
            'fixed_assets': {
                'tangible_assets': {'entries': [], 'total': 0},
                'intangible_assets': {'entries': [], 'total': 0},
                'total': 0
            },
            'investments': {
                'long_term_investments': {'entries': [], 'total': 0},
                'short_term_investments': {'entries': [], 'total': 0},
                'total': 0
            },
            'total': 0
        },
        'liabilities': {
            'equity': {
                'equity_share_capital': {'entries': [], 'total': 0},
                'general_reserve': {'entries': [], 'total': 0},
                'capital_reserve': {'entries': [], 'total': 0},
                'securities_premium': {'entries': [], 'total': 0},
                'retained_earnings': {'entries': [], 'total': 0},
                'total': 0
            },
            'long_term_liabilities': {
                'secured_loans': {'entries': [], 'total': 0},
                'unsecured_loans': {'entries': [], 'total': 0},
                'total': 0
            },
            'current_liabilities': {
                'msme_payables': {'entries': [], 'total': 0},
                'other_payables': {'entries': [], 'total': 0},
                'statutory_dues': {'entries': [], 'total': 0},
                'short_term_borrowings': {'entries': [], 'total': 0},
                'other_current_liabilities': {'entries': [], 'total': 0},
                'total': 0
            },
            'total': 0
        },
        'total_assets': 0,
        'total_liabilities': 0
    }


def generate_balance_sheet(client_id, trial_balance_data, year_end_date):
    """Generate balance sheet with proper structure and current year profit"""
    try:
        # First, calculate profit/loss for the year
        current_profit_loss = generate_profit_loss_statement(trial_balance_data)
        if current_profit_loss:
            current_year_profit = current_profit_loss.get('profit_for_year', 0)
        else:
            current_year_profit = 0
            
        print(f"Current Year Profit Calculated: {current_year_profit}")

        # Initialize empty balance sheet structure
        balance_sheet = {
            'assets': {
                'current_assets': {
                    'cash_and_equivalents': {'entries': [], 'total': 0},
                    'trade_receivables': {'entries': [], 'total': 0},
                    'inventories': {'entries': [], 'total': 0},
                    'loans_and_advances': {'entries': [], 'total': 0},
                    'other_current_assets': {'entries': [], 'total': 0},
                    'total': 0
                },
                'fixed_assets': {
                    'tangible_assets': {'entries': [], 'total': 0},
                    'intangible_assets': {'entries': [], 'total': 0},
                    'total': 0
                },
                'investments': {
                    'long_term_investments': {'entries': [], 'total': 0},
                    'short_term_investments': {'entries': [], 'total': 0},
                    'total': 0
                },
                'total': 0
            },
            'liabilities': {
                'equity': {
                    'equity_share_capital': {'entries': [], 'total': 0},
                    'general_reserve': {'entries': [], 'total': 0},
                    'capital_reserve': {'entries': [], 'total': 0},
                    'securities_premium': {'entries': [], 'total': 0},
                    'retained_earnings': {
                        'entries': [],
                        'opening_balance': 0,
                        'current_year_profit': current_year_profit,
                        'total': 0
                    },
                    'total': 0
                },
                'long_term_liabilities': {
                    'secured_loans': {'entries': [], 'total': 0},
                    'unsecured_loans': {'entries': [], 'total': 0},
                    'total': 0
                },
                'current_liabilities': {
                    'msme_payables': {'entries': [], 'total': 0},
                    'other_payables': {'entries': [], 'total': 0},
                    'statutory_dues': {'entries': [], 'total': 0},
                    'short_term_borrowings': {'entries': [], 'total': 0},
                    'other_current_liabilities': {'entries': [], 'total': 0},
                    'total': 0
                },
                'total': 0
            },
            'total_assets': 0,
            'total_liabilities': 0
        }

        # Process trial balance entries
        retained_earnings_total = 0
        for entry in trial_balance_data:
            try:
                debit = float(entry.get('Debit', 0) or 0)
                credit = float(entry.get('Credit', 0) or 0)
                mapping = str(entry.get('Mapping', '')).strip()
                account_name = str(entry.get('Account_Name', '')).strip()
                account_type = str(entry.get('Account_Type', '')).strip()

                # Skip P&L entries as they're reflected in current_year_profit
                if mapping.startswith('PL-'):
                    continue

                amount = credit - debit  # Default for liabilities and equity
                if account_type == 'Asset':
                    amount = debit - credit  # For assets, use debit - credit

                if mapping.startswith('BS-'):
                    subschedule = mapping[3:]  # Remove 'BS-' prefix
                    
                    # Current Assets
                    if subschedule == 'cash_and_equivalents':
                        balance_sheet['assets']['current_assets']['cash_and_equivalents']['entries'].append({
                            'account': account_name, 'amount': amount
                        })
                        balance_sheet['assets']['current_assets']['cash_and_equivalents']['total'] += amount
                    
                    elif subschedule == 'trade_receivables':
                        balance_sheet['assets']['current_assets']['trade_receivables']['entries'].append({
                            'account': account_name, 'amount': amount
                        })
                        balance_sheet['assets']['current_assets']['trade_receivables']['total'] += amount
                    
                    elif subschedule == 'inventories':
                        balance_sheet['assets']['current_assets']['inventories']['entries'].append({
                            'account': account_name, 'amount': amount
                        })
                        balance_sheet['assets']['current_assets']['inventories']['total'] += amount
                    
                    elif subschedule == 'loans_and_advances':
                        balance_sheet['assets']['current_assets']['loans_and_advances']['entries'].append({
                            'account': account_name, 'amount': amount
                        })
                        balance_sheet['assets']['current_assets']['loans_and_advances']['total'] += amount
                    
                    elif subschedule == 'other_current_assets':
                        balance_sheet['assets']['current_assets']['other_current_assets']['entries'].append({
                            'account': account_name, 'amount': amount
                        })
                        balance_sheet['assets']['current_assets']['other_current_assets']['total'] += amount
                    
                    # Fixed Assets
                    elif subschedule == 'tangible_assets':
                        balance_sheet['assets']['fixed_assets']['tangible_assets']['entries'].append({
                            'account': account_name, 'amount': amount
                        })
                        balance_sheet['assets']['fixed_assets']['tangible_assets']['total'] += amount
                    
                    elif subschedule == 'intangible_assets':
                        balance_sheet['assets']['fixed_assets']['intangible_assets']['entries'].append({
                            'account': account_name, 'amount': amount
                        })
                        balance_sheet['assets']['fixed_assets']['intangible_assets']['total'] += amount
                    
                    # Investments
                    elif subschedule == 'long_term_investments':
                        balance_sheet['assets']['investments']['long_term_investments']['entries'].append({
                            'account': account_name, 'amount': amount
                        })
                        balance_sheet['assets']['investments']['long_term_investments']['total'] += amount
                    
                    elif subschedule == 'short_term_investments':
                        balance_sheet['assets']['investments']['short_term_investments']['entries'].append({
                            'account': account_name, 'amount': amount
                        })
                        balance_sheet['assets']['investments']['short_term_investments']['total'] += amount
                    
                    # Equity
                    elif subschedule == 'equity_share_capital':
                        balance_sheet['liabilities']['equity']['equity_share_capital']['entries'].append({
                            'account': account_name, 'amount': amount
                        })
                        balance_sheet['liabilities']['equity']['equity_share_capital']['total'] += amount
                    
                    elif subschedule == 'general_reserve':
                        balance_sheet['liabilities']['equity']['general_reserve']['entries'].append({
                            'account': account_name, 'amount': amount
                        })
                        balance_sheet['liabilities']['equity']['general_reserve']['total'] += amount
                    
                    elif subschedule == 'capital_reserve':
                        balance_sheet['liabilities']['equity']['capital_reserve']['entries'].append({
                            'account': account_name, 'amount': amount
                        })
                        balance_sheet['liabilities']['equity']['capital_reserve']['total'] += amount
                    
                    elif subschedule == 'securities_premium':
                        balance_sheet['liabilities']['equity']['securities_premium']['entries'].append({
                            'account': account_name, 'amount': amount
                        })
                        balance_sheet['liabilities']['equity']['securities_premium']['total'] += amount
                    
                    elif subschedule == 'retained_earnings':
                        retained_earnings_total += amount
                        balance_sheet['liabilities']['equity']['retained_earnings']['entries'].append({
                            'account': account_name,
                            'amount': amount
                        })
                    
                    # Long Term Liabilities
                    elif subschedule == 'secured_loans':
                        balance_sheet['liabilities']['long_term_liabilities']['secured_loans']['entries'].append({
                            'account': account_name, 'amount': amount
                        })
                        balance_sheet['liabilities']['long_term_liabilities']['secured_loans']['total'] += amount
                    
                    elif subschedule == 'unsecured_loans':
                        balance_sheet['liabilities']['long_term_liabilities']['unsecured_loans']['entries'].append({
                            'account': account_name, 'amount': amount
                        })
                        balance_sheet['liabilities']['long_term_liabilities']['unsecured_loans']['total'] += amount
                    
                    # Current Liabilities
                    elif subschedule == 'msme_payables':
                        balance_sheet['liabilities']['current_liabilities']['msme_payables']['entries'].append({
                            'account': account_name, 'amount': amount
                        })
                        balance_sheet['liabilities']['current_liabilities']['msme_payables']['total'] += amount
                    
                    elif subschedule == 'other_payables':
                        balance_sheet['liabilities']['current_liabilities']['other_payables']['entries'].append({
                            'account': account_name, 'amount': amount
                        })
                        balance_sheet['liabilities']['current_liabilities']['other_payables']['total'] += amount
                    
                    elif subschedule == 'statutory_dues':
                        balance_sheet['liabilities']['current_liabilities']['statutory_dues']['entries'].append({
                            'account': account_name, 'amount': amount
                        })
                        balance_sheet['liabilities']['current_liabilities']['statutory_dues']['total'] += amount
                    
                    elif subschedule == 'short_term_borrowings':
                        balance_sheet['liabilities']['current_liabilities']['short_term_borrowings']['entries'].append({
                            'account': account_name, 'amount': amount
                        })
                        balance_sheet['liabilities']['current_liabilities']['short_term_borrowings']['total'] += amount
                    
                    elif subschedule == 'other_current_liabilities':
                        balance_sheet['liabilities']['current_liabilities']['other_current_liabilities']['entries'].append({
                            'account': account_name, 'amount': amount
                        })
                        balance_sheet['liabilities']['current_liabilities']['other_current_liabilities']['total'] += amount

            except Exception as e:
                print(f"Error processing entry: {str(e)}")
                continue

        # Set opening balance of retained earnings
        balance_sheet['liabilities']['equity']['retained_earnings']['opening_balance'] = retained_earnings_total

        # Calculate total retained earnings including current year profit
        balance_sheet['liabilities']['equity']['retained_earnings']['total'] = (
            retained_earnings_total + current_year_profit
        )

        # Calculate section totals
        # Current Assets Total
        balance_sheet['assets']['current_assets']['total'] = (
            balance_sheet['assets']['current_assets']['cash_and_equivalents']['total'] +
            balance_sheet['assets']['current_assets']['trade_receivables']['total'] +
            balance_sheet['assets']['current_assets']['inventories']['total'] +
            balance_sheet['assets']['current_assets']['loans_and_advances']['total'] +
            balance_sheet['assets']['current_assets']['other_current_assets']['total']
        )

        # Fixed Assets Total
        balance_sheet['assets']['fixed_assets']['total'] = (
            balance_sheet['assets']['fixed_assets']['tangible_assets']['total'] +
            balance_sheet['assets']['fixed_assets']['intangible_assets']['total']
        )

        # Investments Total
        balance_sheet['assets']['investments']['total'] = (
            balance_sheet['assets']['investments']['long_term_investments']['total'] +
            balance_sheet['assets']['investments']['short_term_investments']['total']
        )

        # Total Assets
        balance_sheet['assets']['total'] = (
            balance_sheet['assets']['current_assets']['total'] +
            balance_sheet['assets']['fixed_assets']['total'] +
            balance_sheet['assets']['investments']['total']
        )

        # Equity Total
        balance_sheet['liabilities']['equity']['total'] = (
            balance_sheet['liabilities']['equity']['equity_share_capital']['total'] +
            balance_sheet['liabilities']['equity']['general_reserve']['total'] +
            balance_sheet['liabilities']['equity']['capital_reserve']['total'] +
            balance_sheet['liabilities']['equity']['securities_premium']['total'] +
            balance_sheet['liabilities']['equity']['retained_earnings']['total']
        )

        # Long Term Liabilities Total
        balance_sheet['liabilities']['long_term_liabilities']['total'] = (
            balance_sheet['liabilities']['long_term_liabilities']['secured_loans']['total'] +
            balance_sheet['liabilities']['long_term_liabilities']['unsecured_loans']['total']
        )

        # Current Liabilities Total
        balance_sheet['liabilities']['current_liabilities']['total'] = (
            balance_sheet['liabilities']['current_liabilities']['msme_payables']['total'] +
            balance_sheet['liabilities']['current_liabilities']['other_payables']['total'] +
            balance_sheet['liabilities']['current_liabilities']['statutory_dues']['total'] +
            balance_sheet['liabilities']['current_liabilities']['short_term_borrowings']['total'] +
            balance_sheet['liabilities']['current_liabilities']['other_current_liabilities']['total']
        )

        # Total Liabilities
        balance_sheet['liabilities']['total'] = (
            balance_sheet['liabilities']['equity']['total'] +
            balance_sheet['liabilities']['long_term_liabilities']['total'] +
            balance_sheet['liabilities']['current_liabilities']['total']
        )

        balance_sheet['total_assets'] = balance_sheet['assets']['total']
        balance_sheet['total_liabilities'] = balance_sheet['liabilities']['total']

        # Verify balance sheet equation
        if not math.isclose(balance_sheet['total_assets'], balance_sheet['total_liabilities'], rel_tol=1e-9):
            print(f"Warning: Balance sheet not balanced!")
            print(f"Total Assets: {balance_sheet['total_assets']}")
            print(f"Total Liabilities: {balance_sheet['total_liabilities']}")
            print(f"Difference: {balance_sheet['total_assets'] - balance_sheet['total_liabilities']}")

        return balance_sheet

    except Exception as e:
        print(f"Error generating balance sheet: {str(e)}")
        print(f"Trial balance data: {trial_balance_data}")
        raise

@financials.route('/generate_balance_sheet/<client_id>')
@login_required
def generate_balance_sheet_view(client_id):
    try:
        db = get_db()
        
        # Debug logging
        print(f"\nStarting balance sheet generation for client_id: {client_id}")
        
        # Get client info
        client = db.financial_clients.find_one({
            '_id': ObjectId(client_id),
            'user_id': session['user_id']
        })
        
        if not client:
            flash('Client not found', 'error')
            return redirect(url_for('financials.financial_dashboard'))
            
        # Get trial balance data
        trial_balance = db.trial_balances.find_one({
            'client_id': ObjectId(client_id),
            'user_id': session['user_id']
        })
        
        if not trial_balance:
            flash('Trial balance not found. Please upload trial balance first.', 'error')
            return redirect(url_for('financials.view_client_details', client_id=client_id))
            
        print(f"Found trial balance data: {len(trial_balance.get('data', []))} entries")
        
        # Continue with the rest of your code...
            
        # Calculate balance sheet dates
        balance_sheet_date = datetime.strptime(client['end_date'], '%Y-%m-%d') if isinstance(client['end_date'], str) else client['end_date']
        previous_year_date = balance_sheet_date - timedelta(days=365)
        
        # Get previous year trial balance
        previous_trial_balance = db.trial_balances.find_one({
            'client_id': ObjectId(client_id),
            'user_id': session['user_id'],
            'period.end_date': previous_year_date.strftime('%Y-%m-%d')
        })
        
        # Generate current year balance sheet
        current_balance_sheet = generate_balance_sheet(
            client_id, 
            trial_balance['data'],
            balance_sheet_date
        )
        
        # Generate previous year balance sheet (or empty structure if no data)
        if previous_trial_balance:
            previous_balance_sheet = generate_balance_sheet(
                client_id,
                previous_trial_balance['data'],
                previous_year_date
            )
        else:
            previous_balance_sheet = get_empty_balance_sheet()
        
        # Calculate totals
        # Calculate totals
        current_balance_sheet['total_liabilities'] = (
            current_balance_sheet['liabilities']['equity']['total'] +
            current_balance_sheet['liabilities']['long_term_liabilities']['total'] +
            current_balance_sheet['liabilities']['current_liabilities']['total']
        )

        current_balance_sheet['total_assets'] = current_balance_sheet['assets']['total']

        if previous_balance_sheet:
            previous_balance_sheet['total_liabilities'] = (
                previous_balance_sheet['liabilities']['equity']['total'] +
                previous_balance_sheet['liabilities']['long_term_liabilities']['total'] +
                previous_balance_sheet['liabilities']['current_liabilities']['total']
            )
            previous_balance_sheet['total_assets'] = previous_balance_sheet['assets']['total']        
        return render_template(
            'financials/balance_sheet.html',
            client=client,
            balance_sheet=current_balance_sheet,
            previous_balance_sheet=previous_balance_sheet,
            balance_sheet_date=balance_sheet_date,
            previous_year_date=previous_year_date,
            format_amount=lambda x: format_amount(x, client['settings']['rounding_amount'], int(client['settings']['decimal_rounding'])),
            format_date=lambda d: d.strftime('%B %d, %Y') if isinstance(d, datetime) else d
        )
        
    except Exception as e:
        logging.error(f"Error generating balance sheet: {str(e)}")
        logging.error(traceback.format_exc())
        flash(f'Error generating balance sheet: {str(e)}', 'error')
        return redirect(url_for('financials.view_client_details', client_id=client_id))



@financials.route('/export_balance_sheet/<client_id>')
@login_required
def export_balance_sheet(client_id):
    try:
        db = get_db()
        
        # Get client and balance sheet data (reuse logic from generate_balance_sheet_view)
        client = db.financial_clients.find_one({
            '_id': ObjectId(client_id),
            'user_id': session['user_id']
        })
        
        if not client:
            flash('Client not found', 'error')
            return redirect(url_for('financials.financial_dashboard'))
            
        # Create Excel workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Balance Sheet"
        
        # Add company header
        ws['A1'] = client['legal_name']
        ws['A2'] = 'Balance Sheet'
        ws['A3'] = f'As at {client["end_date"]}'
        ws['A4'] = f'All amounts in INR {client["settings"]["rounding_amount"]}, unless otherwise stated'
        
        # Add header formatting
        for row in range(1, 5):
            cell = ws[f'A{row}']
            cell.font = Font(bold=True)
            if row == 1:
                cell.font.size = 14
            ws.merge_cells(f'A{row}:D{row}')
            cell.alignment = Alignment(horizontal='center')
            
        # Add headers
        headers = ['Particulars', 'Note', f'As at {client["end_date"]}', f'As at {previous_year_date}']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=6, column=col)
            cell.value = header
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="E6E6E6", end_color="E6E6E6", fill_type="solid")
            
        # Add data
        # ... (Add balance sheet data to Excel)
        
        # Save to BytesIO
        excel_file = BytesIO()
        wb.save(excel_file)
        excel_file.seek(0)
        
        return send_file(
            excel_file,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=f'Balance_Sheet_{client["legal_name"]}_{client["end_date"]}.xlsx'
        )
        
    except Exception as e:
        logging.error(f"Error exporting balance sheet: {str(e)}")
        flash(f'Error exporting balance sheet: {str(e)}', 'error')
        return redirect(url_for('financials.generate_balance_sheet_view', client_id=client_id))

@financials.route('/generate_profit_loss/<client_id>')
@login_required
def generate_profit_loss(client_id):
    try:
        db = get_db()
        
        # Get client info
        client = db.financial_clients.find_one({
            '_id': ObjectId(client_id),
            'user_id': session['user_id']
        })
        
        if not client:
            flash('Client not found', 'error')
            return redirect(url_for('financials.financial_dashboard'))
            
        # Get trial balance data
        trial_balance = db.trial_balances.find_one({
            'client_id': ObjectId(client_id),
            'user_id': session['user_id']
        })
        
        if not trial_balance:
            flash('Trial balance not found. Please upload trial balance first.', 'error')
            return redirect(url_for('financials.view_client_details', client_id=client_id))

        # Process trial balance to generate P&L statement
        profit_loss_data = generate_profit_loss_statement(trial_balance['data'])
        
        return render_template(
            'financials/profit_loss.html',
            client=client,
            profit_loss=profit_loss_data,
            format_amount=lambda x: format_amount(
                x, 
                client['settings']['rounding_amount'], 
                client['settings'].get('decimal_rounding', 2)
            )
        )
        
    except Exception as e:
        logging.error(f"Error generating profit and loss: {str(e)}")
        flash(f'Error generating profit and loss statement: {str(e)}', 'error')
        return redirect(url_for('financials.view_client_details', client_id=client_id))

def generate_profit_loss_statement(trial_balance_data):
    """Generate profit and loss statement from trial balance data with proper type handling"""
    try:
        # Initialize profit loss structure
        profit_loss = {
            'income': {
                'revenue_from_operations': 0,
                'other_income': 0,
                'total_income': 0
            },
            'expenses': {
                'cost_of_materials': 0,
                'employee_benefits': 0,
                'finance_costs': 0,
                'depreciation': 0,
                'other_expenses': 0,
                'total_expenses': 0
            },
            'profit_before_tax': 0,
            'tax_expenses': {
                'current_tax': 0,
                'total_tax': 0
            },
            'profit_for_year': 0
        }
        
        # Process trial balance entries
        for entry in trial_balance_data:
            try:
                # Convert debit/credit to float and handle None values
                debit = float(entry.get('Debit', 0) or 0)
                credit = float(entry.get('Credit', 0) or 0)
                mapping = str(entry.get('Mapping', '')).strip()
                
                if mapping.startswith('PL-'):
                    main_category = get_main_pl_category(mapping)
                    section = get_pl_section(mapping)
                    
                    if not main_category or not section:
                        continue
                        
                    # Calculate amount based on section
                    if section == 'income':
                        amount = credit - debit  # Income items are credit minus debit
                    else:
                        amount = debit - credit  # Expense items are debit minus credit
                        
                    # Add to appropriate category
                    if section == 'income':
                        if main_category == 'revenue_from_operations':
                            profit_loss['income']['revenue_from_operations'] += amount
                        elif main_category == 'other_income':
                            profit_loss['income']['other_income'] += amount
                    elif section == 'expenses':
                        if main_category == 'cost_of_materials':
                            profit_loss['expenses']['cost_of_materials'] += amount
                        elif main_category == 'employee_benefits':
                            profit_loss['expenses']['employee_benefits'] += amount
                        elif main_category == 'finance_costs':
                            profit_loss['expenses']['finance_costs'] += amount
                        elif main_category == 'depreciation_and_amortisation':
                            profit_loss['expenses']['depreciation'] += amount
                        elif main_category == 'other_expenses':
                            profit_loss['expenses']['other_expenses'] += amount
                        elif main_category == 'tax_expenses':
                            profit_loss['tax_expenses']['current_tax'] += amount
                
            except Exception as e:
                print(f"Error processing entry {entry}: {str(e)}")
                continue

        # Calculate totals
        profit_loss['income']['total_income'] = (
            profit_loss['income']['revenue_from_operations'] +
            profit_loss['income']['other_income']
        )
        
        profit_loss['expenses']['total_expenses'] = (
            profit_loss['expenses']['cost_of_materials'] +
            profit_loss['expenses']['employee_benefits'] +
            profit_loss['expenses']['finance_costs'] +
            profit_loss['expenses']['depreciation'] +
            profit_loss['expenses']['other_expenses']
        )
        
        profit_loss['profit_before_tax'] = (
            profit_loss['income']['total_income'] -
            profit_loss['expenses']['total_expenses']
        )
        
        profit_loss['tax_expenses']['total_tax'] = profit_loss['tax_expenses']['current_tax']
        
        profit_loss['profit_for_year'] = (
            profit_loss['profit_before_tax'] -
            profit_loss['tax_expenses']['total_tax']
        )

        # Debug logging
        print("\nProfit & Loss Calculation Results:")
        print(f"Total Income: {profit_loss['income']['total_income']}")
        print(f"Total Expenses: {profit_loss['expenses']['total_expenses']}")
        print(f"Profit Before Tax: {profit_loss['profit_before_tax']}")
        print(f"Tax: {profit_loss['tax_expenses']['total_tax']}")
        print(f"Final Profit/Loss: {profit_loss['profit_for_year']}")
        
        return profit_loss
        
    except Exception as e:
        print(f"Error generating profit and loss statement: {str(e)}")
        print(f"Trial balance data: {trial_balance_data}")
        raise


@financials.route('/export_profit_loss_excel/<client_id>')
@login_required
def export_profit_loss_excel(client_id):
    try:
        db = get_db()
        client = db.financial_clients.find_one({"_id": ObjectId(client_id)})
        
        if not client:
            flash('Client not found', 'error')
            return redirect(url_for('financials.financial_dashboard'))
            
        # Get trial balance data
        trial_balance = db.trial_balances.find_one({
            "client_id": ObjectId(client_id),
            "user_id": session['user_id']
        })
        
        if not trial_balance:
            flash('Trial balance not found', 'error')
            return redirect(url_for('financials.view_client_details', client_id=client_id))

        # Process the data
        profit_loss_data = generate_profit_loss_statement(trial_balance['data'])
        
        # Create Excel workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Profit and Loss"
        
        # Add headers
        ws['A1'] = client['legal_name']
        ws['A2'] = 'Statement of Profit And Loss'
        ws['A3'] = f'Year ended March 31, {client["fiscal_year"]}'
        ws['A4'] = f'All amounts in INR {client["settings"]["rounding_amount"]}, unless otherwise stated'
        
        # Format headers
        for i in range(1, 5):
            cell = ws[f'A{i}']
            cell.font = Font(bold=True)
            if i == 1:
                cell.font.size = 14
            ws.merge_cells(f'A{i}:D{i}')
            cell.alignment = Alignment(horizontal='center')

        # Add column headers
        headers = ['Particulars', 'Note', f'Year ended March 31, {client["fiscal_year"]}', 
                  f'Year ended March 31, {int(client["fiscal_year"]) - 1}']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=6, column=col)
            cell.value = header
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="E6E6E6", end_color="E6E6E6", fill_type="solid")

        current_row = 7

        # Income Section
        ws.cell(row=current_row, column=1, value="INCOME").font = Font(bold=True)
        current_row += 1

        income_data = [
            ('Revenue From Operations', '12', profit_loss_data['income']['revenue_from_operations']),
            ('Other Income', '13', profit_loss_data['income']['other_income']),
            ('Total Income', '', profit_loss_data['income']['total_income'])
        ]

        for item in income_data:
            row = [item[0], item[1], item[2]]
            for col, value in enumerate(row, 1):
                cell = ws.cell(row=current_row, column=col)
                cell.value = value
                if item[0] == 'Total Income':
                    cell.font = Font(bold=True)
            current_row += 1

        # Expenses Section
        current_row += 1
        ws.cell(row=current_row, column=1, value="EXPENSES").font = Font(bold=True)
        current_row += 1

        expense_data = [
            ('Cost of Materials Consumed', '14', profit_loss_data['expenses']['cost_of_materials']),
            ('Changes in inventories', '15', profit_loss_data['expenses']['changes_in_inventory']),
            ('Employee Benefits Expense', '-', profit_loss_data['expenses']['employee_benefits']),
            ('Depreciation and Amortisation Expense', '-', profit_loss_data['expenses']['depreciation']),
            ('Other Expenses', '16', profit_loss_data['expenses']['other_expenses']),
            ('Total Expenses', '', profit_loss_data['expenses']['total_expenses'])
        ]

        for item in expense_data:
            row = [item[0], item[1], item[2]]
            for col, value in enumerate(row, 1):
                cell = ws.cell(row=current_row, column=col)
                cell.value = value
                if item[0] == 'Total Expenses':
                    cell.font = Font(bold=True)
            current_row += 1

        # Profit/Tax Section
        current_row += 1
        profit_tax_data = [
            ('Profit before tax', '', profit_loss_data['profit_before_tax']),
            ('TAX EXPENSES', '', ''),
            ('Current Tax', '-', profit_loss_data['tax_expenses']['current_tax']),
            ('PROFIT FOR THE YEAR', '', profit_loss_data['profit_for_year'])
        ]

        for item in profit_tax_data:
            row = [item[0], item[1], item[2]]
            for col, value in enumerate(row, 1):
                cell = ws.cell(row=current_row, column=col)
                cell.value = value
                if item[0] in ['Profit before tax', 'TAX EXPENSES', 'PROFIT FOR THE YEAR']:
                    cell.font = Font(bold=True)
            current_row += 1

        # Adjust column widths
        for col in range(1, 5):
            ws.column_dimensions[get_column_letter(col)].width = 25

        # Create response
        excel_file = BytesIO()
        wb.save(excel_file)
        excel_file.seek(0)

        return send_file(
            excel_file,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=f'Profit_and_Loss_{client["legal_name"]}_{client["fiscal_year"]}.xlsx'
        )

    except Exception as e:
        logging.error(f"Error exporting profit and loss to Excel: {str(e)}")
        flash(f'Error exporting to Excel: {str(e)}', 'error')
        return redirect(url_for('financials.generate_profit_loss', client_id=client_id))

@financials.route('/generate_cash_flow/<client_id>')
@login_required
def generate_cash_flow(client_id):
    try:
        db = get_db()
        
        client = db.financial_clients.find_one({
            '_id': ObjectId(client_id),
            'user_id': session['user_id']
        })
        
        if not client:
            flash('Client not found', 'error')
            return redirect(url_for('financials.financial_dashboard'))

        current_trial_balance = db.trial_balances.find_one({
            'client_id': ObjectId(client_id),
            'user_id': session['user_id']
        })
        
        if not current_trial_balance:
            flash('Trial balance not found', 'error')
            return redirect(url_for('financials.view_client_details', client_id=client_id))

        # Calculate dates
        balance_sheet_date = datetime.strptime(client['end_date'], '%Y-%m-%d') if isinstance(client['end_date'], str) else client['end_date']
        previous_year_date = balance_sheet_date - timedelta(days=365)
        
        # Get previous year trial balance
        previous_trial_balance = db.trial_balances.find_one({
            'client_id': ObjectId(client_id),
            'user_id': session['user_id'],
            'period.end_date': previous_year_date.strftime('%Y-%m-%d')
        })

        # Generate required statements
        current_balance_sheet = generate_balance_sheet(
            client_id, 
            current_trial_balance['data'],
            balance_sheet_date
        )
        
        previous_balance_sheet = generate_balance_sheet(
            client_id,
            previous_trial_balance['data'] if previous_trial_balance else [],
            previous_year_date
        )

        profit_loss_data = generate_profit_loss_statement(current_trial_balance['data'])

        cash_flow_data = generate_cash_flow_statement(
            current_balance_sheet,
            previous_balance_sheet,
            profit_loss_data
        )

        # Helper function for safe dictionary access
        def safe_get(dict_obj, path, default=0):
            try:
                value = dict_obj
                for key in path.split('.'):
                    value = value.get(key, {})
                return value if value not in (None, {}) else default
            except:
                return default

        return render_template(
            'financials/cash_flow.html',
            client=client,
            cash_flow=cash_flow_data,
            format_amount=lambda x: format_amount(
                x, 
                client['settings']['rounding_amount'], 
                int(client['settings'].get('decimal_rounding', 2))
            ),
            safe_get=safe_get
        )
        
    except Exception as e:
        print(f"Error generating cash flow: {str(e)}")
        print(traceback.format_exc())
        flash('Error generating cash flow statement', 'error')
        return redirect(url_for('financials.view_client_details', client_id=client_id))

def generate_cash_flow_statement(current_balance_sheet, previous_balance_sheet, profit_loss_data):
    """
    Generate a complete cash flow statement with all components and proper error handling.
    
    Args:
        current_balance_sheet (dict): Current year balance sheet data
        previous_balance_sheet (dict): Previous year balance sheet data
        profit_loss_data (dict): Current year profit and loss data
        
    Returns:
        dict: Complete cash flow statement with all sections and comparatives
    """
    try:
        # Initialize the cash flow structure
        cash_flow = {
            'operating_activities': {
                'profit_before_tax': profit_loss_data.get('profit_before_tax', 0),
                'prev_profit_before_tax': 0,
                'adjustments': {
                    'depreciation': profit_loss_data.get('expenses', {}).get('depreciation', 0),
                    'prev_depreciation': 0,
                    'interest_expense': profit_loss_data.get('expenses', {}).get('finance_costs', 0),
                    'prev_interest_expense': 0,
                    'interest_income': profit_loss_data.get('income', {}).get('other_income', 0),
                    'prev_interest_income': 0,
                    'fixed_assets_profit_loss': 0,
                    'prev_fixed_assets_profit_loss': 0,
                    'total': 0,
                    'prev_total': 0
                },
                'operating_profit_before_wc': 0,
                'prev_operating_profit_before_wc': 0,
                'working_capital': {
                    'inventories': 0,
                    'prev_inventories': 0,
                    'trade_receivables': 0,
                    'prev_trade_receivables': 0,
                    'other_current_assets': 0,
                    'prev_other_current_assets': 0,
                    'trade_payables': 0,
                    'prev_trade_payables': 0,
                    'other_current_liabilities': 0,
                    'prev_other_current_liabilities': 0,
                    'total': 0,
                    'prev_total': 0
                },
                'cash_generated': 0,
                'prev_cash_generated': 0,
                'direct_taxes': profit_loss_data.get('tax_expenses', {}).get('current_tax', 0),
                'prev_direct_taxes': 0,
                'net_cash': 0,
                'prev_net_cash': 0
            },
            'investing_activities': {
                'fixed_assets_purchase': 0,
                'prev_fixed_assets_purchase': 0,
                'fixed_assets_sale': 0,
                'prev_fixed_assets_sale': 0,
                'investments_made': 0,
                'prev_investments_made': 0,
                'investments_sale': 0,
                'prev_investments_sale': 0,
                'interest_received': profit_loss_data.get('income', {}).get('interest_income', 0),
                'prev_interest_received': 0,
                'dividend_received': 0,
                'prev_dividend_received': 0,
                'net_cash': 0,
                'prev_net_cash': 0
            },
            'financing_activities': {
                'share_capital_proceeds': 0,
                'prev_share_capital_proceeds': 0,
                'long_term_borrowings': 0,
                'prev_long_term_borrowings': 0,
                'long_term_borrowings_repayment': 0,
                'prev_long_term_borrowings_repayment': 0,
                'short_term_borrowings_net': 0,
                'prev_short_term_borrowings_net': 0,
                'interest_paid': profit_loss_data.get('expenses', {}).get('finance_costs', 0),
                'prev_interest_paid': 0,
                'dividend_paid': 0,
                'prev_dividend_paid': 0,
                'net_cash': 0,
                'prev_net_cash': 0
            },
            'net_change': 0,
            'prev_net_change': 0,
            'cash_beginning': 0,
            'prev_cash_beginning': 0,
            'cash_ending': 0,
            'prev_cash_ending': 0,
            'components': {
                'cash_on_hand': 0,
                'prev_cash_on_hand': 0,
                'current_accounts': 0,
                'prev_current_accounts': 0,
                'deposit_accounts': 0,
                'prev_deposit_accounts': 0,
                'total': 0,
                'prev_total': 0
            }
        }

        # Helper function to safely get balance sheet values

        # 1. Calculate Operating Activities
        # Add all adjustments
        cash_flow['operating_activities']['adjustments']['total'] = (
            cash_flow['operating_activities']['adjustments']['depreciation'] +
            cash_flow['operating_activities']['adjustments']['interest_expense'] -
            cash_flow['operating_activities']['adjustments']['interest_income'] +
            cash_flow['operating_activities']['adjustments']['fixed_assets_profit_loss']
        )

        # Calculate operating profit before working capital changes
        cash_flow['operating_activities']['operating_profit_before_wc'] = (
            cash_flow['operating_activities']['profit_before_tax'] +
            cash_flow['operating_activities']['adjustments']['total']
        )

        # Calculate working capital changes
        # Assets (increase is negative cash flow)
        for key in ['inventories', 'trade_receivables', 'other_current_assets']:
            current = get_bs_value(current_balance_sheet, f'assets.current_assets.{key}')
            previous = get_bs_value(previous_balance_sheet, f'assets.current_assets.{key}')
            cash_flow['operating_activities']['working_capital'][key] = -(current - previous)

        # Calculate trade payables change
        current_trade_payables = (
            get_bs_value(current_balance_sheet, 'liabilities.current_liabilities.msme_payables.total') +
            get_bs_value(current_balance_sheet, 'liabilities.current_liabilities.other_payables.total')
        )
        
        previous_trade_payables = (
            get_bs_value(previous_balance_sheet, 'liabilities.current_liabilities.msme_payables.total') +
            get_bs_value(previous_balance_sheet, 'liabilities.current_liabilities.other_payables.total')
        )

        # Calculate change - increase in payables is positive (source of cash)
        cash_flow['operating_activities']['working_capital']['trade_payables'] = current_trade_payables - previous_trade_payables
        cash_flow['operating_activities']['working_capital']['prev_trade_payables'] = previous_trade_payables
        # Calculate other current liabilities (includes statutory dues, short term borrowings, and other liabilities)
        current_other_liabilities = (
            get_bs_value(current_balance_sheet, 'liabilities.current_liabilities.statutory_dues.total') +
            get_bs_value(current_balance_sheet, 'liabilities.current_liabilities.short_term_borrowings.total') +
            get_bs_value(current_balance_sheet, 'liabilities.current_liabilities.other_current_liabilities.total')
        )
        
        previous_other_liabilities = (
            get_bs_value(previous_balance_sheet, 'liabilities.current_liabilities.statutory_dues.total') +
            get_bs_value(previous_balance_sheet, 'liabilities.current_liabilities.short_term_borrowings.total') +
            get_bs_value(previous_balance_sheet, 'liabilities.current_liabilities.other_current_liabilities.total')
        )

        # Calculate change - increase in liabilities is positive (source of cash)
        cash_flow['operating_activities']['working_capital']['other_current_liabilities'] = (
            current_other_liabilities - previous_other_liabilities
        )
        cash_flow['operating_activities']['working_capital']['prev_other_current_liabilities'] = previous_other_liabilities

        # Calculate other current assets (including loans and advances and other current assets)
        current_other_assets = (
            get_bs_value(current_balance_sheet, 'assets.current_assets.loans_and_advances.total') +
            get_bs_value(current_balance_sheet, 'assets.current_assets.other_current_assets.total')
        )
        
        previous_other_assets = (
            get_bs_value(previous_balance_sheet, 'assets.current_assets.loans_and_advances.total') +
            get_bs_value(previous_balance_sheet, 'assets.current_assets.other_current_assets.total')
        )

        # Calculate change - for assets, decrease is positive (source of cash)
        # and increase is negative (use of cash)
        cash_flow['operating_activities']['working_capital']['other_current_assets'] = (
            -(current_other_assets - previous_other_assets)  # Note the negative sign
        )
        cash_flow['operating_activities']['working_capital']['prev_other_current_assets'] = previous_other_assets
        
        # Calculate total working capital changes
        cash_flow['operating_activities']['working_capital']['total'] = (
            cash_flow['operating_activities']['working_capital']['trade_payables'] +
            cash_flow['operating_activities']['working_capital']['other_current_liabilities'] +
            cash_flow['operating_activities']['working_capital']['inventories'] +
            cash_flow['operating_activities']['working_capital']['trade_receivables'] +
            cash_flow['operating_activities']['working_capital']['other_current_assets']
        )
        # Calculate cash generated from operations
        cash_flow['operating_activities']['cash_generated'] = (
            cash_flow['operating_activities']['operating_profit_before_wc'] +
            cash_flow['operating_activities']['working_capital']['total']
        )

        # Calculate net operating cash flow
        cash_flow['operating_activities']['net_cash'] = (
            cash_flow['operating_activities']['cash_generated'] -
            cash_flow['operating_activities']['direct_taxes']
        )

        # 2. Calculate Investing Activities
        # Calculate total fixed assets (tangible + intangible) for both periods
        current_fixed_assets = (
            get_bs_value(current_balance_sheet, 'assets.fixed_assets.tangible_assets.total') +
            get_bs_value(current_balance_sheet, 'assets.fixed_assets.intangible_assets.total')
        )
        
        previous_fixed_assets = (
            get_bs_value(previous_balance_sheet, 'assets.fixed_assets.tangible_assets.total') +
            get_bs_value(previous_balance_sheet, 'assets.fixed_assets.intangible_assets.total')
        )

        # Calculate change in fixed assets
        fixed_assets_change = current_fixed_assets - previous_fixed_assets

        # Debug logging
        print("\nFixed Assets Calculation:")
        print("Current Year:")
        print(f"  Tangible Assets: {get_bs_value(current_balance_sheet, 'assets.fixed_assets.tangible_assets.total')}")
        print(f"  Intangible Assets: {get_bs_value(current_balance_sheet, 'assets.fixed_assets.intangible_assets.total')}")
        print(f"  Total: {current_fixed_assets}")
        
        print("\nPrevious Year:")
        print(f"  Tangible Assets: {get_bs_value(previous_balance_sheet, 'assets.fixed_assets.tangible_assets.total')}")
        print(f"  Intangible Assets: {get_bs_value(previous_balance_sheet, 'assets.fixed_assets.intangible_assets.total')}")
        print(f"  Total: {previous_fixed_assets}")
        
        print(f"\nNet Change: {fixed_assets_change}")

        # If there's an increase, it's a purchase (negative cash flow)
        if fixed_assets_change > 0:
            cash_flow['investing_activities']['fixed_assets_purchase'] = fixed_assets_change
            cash_flow['investing_activities']['fixed_assets_sale'] = 0
        else:
            # If there's a decrease, it's a sale (positive cash flow)
            cash_flow['investing_activities']['fixed_assets_purchase'] = 0
            cash_flow['investing_activities']['fixed_assets_sale'] = abs(fixed_assets_change)

        # Investments changes
        current_investments = get_bs_value(current_balance_sheet, 'assets.investments.total')
        previous_investments = get_bs_value(previous_balance_sheet, 'assets.investments.total')
        investment_change = current_investments - previous_investments

        if investment_change > 0:
            cash_flow['investing_activities']['investments_made'] = investment_change
        else:
            cash_flow['investing_activities']['investments_sale'] = abs(investment_change)

                # Get interest income from profit and loss data
        interest_received = get_value(profit_loss_data, 'income.other_income')  # Assuming interest is part of other income
        
        # Debug logging
        print("\nInterest Income Calculation:")
        print(f"Current Year Interest Income: {interest_received}")
        
        # Assign interest income to investing activities
        cash_flow['investing_activities']['interest_received'] = interest_received

        # Calculate net investing cash flow
        cash_flow['investing_activities']['net_cash'] = (
            cash_flow['investing_activities']['fixed_assets_sale'] +
            cash_flow['investing_activities']['investments_sale'] +
            cash_flow['investing_activities']['interest_received'] +
            cash_flow['investing_activities']['dividend_received'] -
            cash_flow['investing_activities']['fixed_assets_purchase'] -
            cash_flow['investing_activities']['investments_made']
        )

        # 3. Calculate Financing Activities
        # Share capital changes
        # Calculate total equity components for current period
        current_equity_components = (
            get_bs_value(current_balance_sheet, 'liabilities.equity.equity_share_capital.total') +
            get_bs_value(current_balance_sheet, 'liabilities.equity.securities_premium.total') +
            get_bs_value(current_balance_sheet, 'liabilities.equity.general_reserve.total') +
            get_bs_value(current_balance_sheet, 'liabilities.equity.capital_reserve.total') +
            get_bs_value(current_balance_sheet, 'liabilities.equity.retained_earnings.opening_balance')
        )
        
        # Calculate total equity components for previous period
        previous_equity_components = (
            get_bs_value(previous_balance_sheet, 'liabilities.equity.equity_share_capital.total') +
            get_bs_value(previous_balance_sheet, 'liabilities.equity.securities_premium.total') +
            get_bs_value(previous_balance_sheet, 'liabilities.equity.general_reserve.total') +
            get_bs_value(previous_balance_sheet, 'liabilities.equity.capital_reserve.total') +
            get_bs_value(previous_balance_sheet, 'liabilities.equity.retained_earnings.opening_balance')
        )

        # Calculate net change in equity
        equity_change = current_equity_components - previous_equity_components

        # Debug logging
        print("\nEquity Components Calculation:")
        print("Current Year Components:")
        print(f"  Share Capital: {get_bs_value(current_balance_sheet, 'liabilities.equity.equity_share_capital.total')}")
        print(f"  Securities Premium: {get_bs_value(current_balance_sheet, 'liabilities.equity.securities_premium.total')}")
        print(f"  General Reserve: {get_bs_value(current_balance_sheet, 'liabilities.equity.general_reserve.total')}")
        print(f"  Capital Reserve: {get_bs_value(current_balance_sheet, 'liabilities.equity.capital_reserve.total')}")
        print(f"  Opening Retained Earnings: {get_bs_value(current_balance_sheet, 'liabilities.equity.retained_earnings.opening_balance')}")
        print(f"  Total Current Equity: {current_equity_components}")
        
        print("\nPrevious Year Components:")
        print(f"  Share Capital: {get_bs_value(previous_balance_sheet, 'liabilities.equity.equity_share_capital.total')}")
        print(f"  Securities Premium: {get_bs_value(previous_balance_sheet, 'liabilities.equity.securities_premium.total')}")
        print(f"  General Reserve: {get_bs_value(previous_balance_sheet, 'liabilities.equity.general_reserve.total')}")
        print(f"  Capital Reserve: {get_bs_value(previous_balance_sheet, 'liabilities.equity.capital_reserve.total')}")
        print(f"  Opening Retained Earnings: {get_bs_value(previous_balance_sheet, 'liabilities.equity.retained_earnings.opening_balance')}")
        print(f"  Total Previous Equity: {previous_equity_components}")
        
        print(f"\nNet Change in Equity Components: {equity_change}")

        # Assign the change to cash flow (positive change means cash inflow)
        cash_flow['financing_activities']['share_capital_proceeds'] = equity_change
        cash_flow['financing_activities']['prev_share_capital_proceeds'] = previous_equity_components

        # Update the description in the template
        cash_flow['financing_activities']['description'] = "Proceeds from Issue of Share Capital and Other Equity Components"

        # Long term borrowings changes
        current_lt_borrowings = get_bs_value(current_balance_sheet, 'liabilities.long_term_liabilities.total')
        previous_lt_borrowings = get_bs_value(previous_balance_sheet, 'liabilities.long_term_liabilities.total')
        borrowings_change = current_lt_borrowings - previous_lt_borrowings

        if borrowings_change > 0:
            cash_flow['financing_activities']['long_term_borrowings'] = borrowings_change
        else:
            cash_flow['financing_activities']['long_term_borrowings_repayment'] = abs(borrowings_change)

        # Calculate net financing cash flow
        cash_flow['financing_activities']['net_cash'] = (
            cash_flow['financing_activities']['share_capital_proceeds'] +
            cash_flow['financing_activities']['long_term_borrowings'] -
            cash_flow['financing_activities']['long_term_borrowings_repayment'] +
            cash_flow['financing_activities']['short_term_borrowings_net'] -
            cash_flow['financing_activities']['interest_paid'] -
            cash_flow['financing_activities']['dividend_paid']
        )

        # 4. Calculate Overall Cash Position
        cash_flow['net_change'] = (
            cash_flow['operating_activities']['net_cash'] +
            cash_flow['investing_activities']['net_cash'] +
            cash_flow['financing_activities']['net_cash']
        )

        # Get cash balances
        cash_flow['cash_beginning'] = get_bs_value(previous_balance_sheet, 'assets.current_assets.cash_and_equivalents')
        cash_flow['cash_ending'] = get_bs_value(current_balance_sheet, 'assets.current_assets.cash_and_equivalents')

        # Calculate cash components
        total_cash = get_bs_value(current_balance_sheet, 'assets.current_assets.cash_and_equivalents')
        cash_flow['components']['cash_on_hand'] = total_cash * 0.1  # Assuming 10% is cash on hand
        cash_flow['components']['current_accounts'] = total_cash * 0.45  # Assuming 45% in current accounts
        cash_flow['components']['deposit_accounts'] = total_cash * 0.45  # Assuming 45% in deposit accounts
        cash_flow['components']['total'] = total_cash

        return cash_flow

    except Exception as e:
        print(f"Error generating cash flow statement: {str(e)}")
        print(f"Current balance sheet: {current_balance_sheet}")
        print(f"Previous balance sheet: {previous_balance_sheet}")
        print(f"Profit/Loss data: {profit_loss_data}")
        traceback.print_exc()
        raise

def get_value(balance_sheet, path):
    """Helper function to safely get nested values from balance sheet"""
    try:
        value = balance_sheet
        for key in path.split('.'):
            value = value.get(key, 0)
        return float(value) if value else 0
    except:
        return 0

def generate_balance_sheet_notes(trial_balance_data, balance_sheet):
    """Generate detailed notes for balance sheet with sub-schedule based mapping"""
    try:
        notes = {
            # Note 3: Share Capital and Reserves
            'note_3': {
                'title': 'Note 3: Shareholders\' Funds',
                'share_capital': {
                    'title': '3.1 Share Capital',
                    'entries': balance_sheet['liabilities']['equity']['equity_share_capital']['entries'],
                    'total': balance_sheet['liabilities']['equity']['equity_share_capital']['total']
                },
                'reserves': {
                    'title': '3.2 Reserves and Surplus',
                    'sub_categories': {
                        'general_reserve': {
                            'title': 'General Reserve',
                            'entries': balance_sheet['liabilities']['equity']['general_reserve']['entries'],
                            'total': balance_sheet['liabilities']['equity']['general_reserve']['total']
                        },
                        'capital_reserve': {
                            'title': 'Capital Reserve',
                            'entries': balance_sheet['liabilities']['equity']['capital_reserve']['entries'],
                            'total': balance_sheet['liabilities']['equity']['capital_reserve']['total']
                        },
                        'securities_premium': {
                            'title': 'Securities Premium',
                            'entries': balance_sheet['liabilities']['equity']['securities_premium']['entries'],
                            'total': balance_sheet['liabilities']['equity']['securities_premium']['total']
                        },
                        'retained_earnings': {
                            'title': 'Retained Earnings',
                            'entries': balance_sheet['liabilities']['equity']['retained_earnings']['entries'],
                            'opening_balance': balance_sheet['liabilities']['equity']['retained_earnings']['opening_balance'],
                            'current_year_profit': balance_sheet['liabilities']['equity']['retained_earnings']['current_year_profit'],
                            'total': balance_sheet['liabilities']['equity']['retained_earnings']['total']
                        }
                    },
                    'total': balance_sheet['liabilities']['equity']['general_reserve']['total'] +
                            balance_sheet['liabilities']['equity']['capital_reserve']['total'] +
                            balance_sheet['liabilities']['equity']['securities_premium']['total'] +
                            balance_sheet['liabilities']['equity']['retained_earnings']['total']
                },
                'total': balance_sheet['liabilities']['equity']['total']
            },

            # Note 4: Long Term Borrowings
            'note_4': {
                'title': 'Note 4: Long Term Borrowings',
                'sub_categories': {
                    'secured_loans': {
                        'title': 'Secured Loans',
                        'entries': balance_sheet['liabilities']['long_term_liabilities']['secured_loans']['entries'],
                        'total': balance_sheet['liabilities']['long_term_liabilities']['secured_loans']['total']
                    },
                    'unsecured_loans': {
                        'title': 'Unsecured Loans',
                        'entries': balance_sheet['liabilities']['long_term_liabilities']['unsecured_loans']['entries'],
                        'total': balance_sheet['liabilities']['long_term_liabilities']['unsecured_loans']['total']
                    }
                },
                'total': balance_sheet['liabilities']['long_term_liabilities']['total']
            },

            # Note 5: Trade Payables
            'note_5': {
                'title': 'Note 5: Trade Payables',
                'sub_categories': {
                    'msme': {
                        'title': 'Due to Micro and Small Enterprises',
                        'entries': balance_sheet['liabilities']['current_liabilities']['msme_payables']['entries'],
                        'total': balance_sheet['liabilities']['current_liabilities']['msme_payables']['total']
                    },
                    'others': {
                        'title': 'Due to Others',
                        'entries': balance_sheet['liabilities']['current_liabilities']['other_payables']['entries'],
                        'total': balance_sheet['liabilities']['current_liabilities']['other_payables']['total']
                    }
                },
                'total': (
                    balance_sheet['liabilities']['current_liabilities']['msme_payables']['total'] +
                    balance_sheet['liabilities']['current_liabilities']['other_payables']['total']
                )
            },

            # Note 6: Other Current Liabilities
            'note_6': {
                'title': 'Note 6: Other Current Liabilities',
                'sub_categories': {
                    'statutory_dues': {
                        'title': 'Statutory Dues',
                        'entries': balance_sheet['liabilities']['current_liabilities']['statutory_dues']['entries'],
                        'total': balance_sheet['liabilities']['current_liabilities']['statutory_dues']['total']
                    },
                    'short_term_borrowings': {
                        'title': 'Short Term Borrowings',
                        'entries': balance_sheet['liabilities']['current_liabilities']['short_term_borrowings']['entries'],
                        'total': balance_sheet['liabilities']['current_liabilities']['short_term_borrowings']['total']
                    },
                    'other_liabilities': {
                        'title': 'Other Liabilities',
                        'entries': balance_sheet['liabilities']['current_liabilities']['other_current_liabilities']['entries'],
                        'total': balance_sheet['liabilities']['current_liabilities']['other_current_liabilities']['total']
                    }
                },
                'total': balance_sheet['liabilities']['current_liabilities']['total']
            },


            # Note 7: Fixed Assets
            'note_7': {
                'title': 'Note 7: Fixed Assets',
                'sub_categories': {
                    'tangible_assets': {
                        'title': 'Tangible Assets',
                        'entries': balance_sheet['assets']['fixed_assets']['tangible_assets']['entries'],
                        'total': balance_sheet['assets']['fixed_assets']['tangible_assets']['total']
                    },
                    'intangible_assets': {
                        'title': 'Intangible Assets',
                        'entries': balance_sheet['assets']['fixed_assets']['intangible_assets']['entries'],
                        'total': balance_sheet['assets']['fixed_assets']['intangible_assets']['total']
                    }
                },
                'total': balance_sheet['assets']['fixed_assets']['total']
            },

# Note 8: Investments
            'note_8': {
                'title': 'Note 8: Investments',
                'entries': balance_sheet['assets']['investments']['long_term_investments']['entries'] + 
                          balance_sheet['assets']['investments']['short_term_investments']['entries'],
                'total': balance_sheet['assets']['investments']['total']
            },

            # Note 9: Inventories
            'note_9': {
                'title': 'Note 9: Inventories',
                'entries': balance_sheet['assets']['current_assets']['inventories']['entries'],
                'total': balance_sheet['assets']['current_assets']['inventories']['total']
            },

            # Note 10: Trade Receivables
            'note_10': {
                'title': 'Note 10: Trade Receivables',
                'entries': balance_sheet['assets']['current_assets']['trade_receivables']['entries'],
                'total': balance_sheet['assets']['current_assets']['trade_receivables']['total']
            },

            # Note 11: Cash and Cash Equivalents
            'note_11': {
                'title': 'Note 11: Cash and Cash Equivalents',
                'entries': balance_sheet['assets']['current_assets']['cash_and_equivalents']['entries'],
                'total': balance_sheet['assets']['current_assets']['cash_and_equivalents']['total']
            },

            # Note 12: Other Current Assets
            'note_12': {
                'title': 'Note 12: Other Current Assets',
                'sub_categories': {
                    'loans_and_advances': {
                        'title': 'Loans and Advances',
                        'entries': balance_sheet['assets']['current_assets']['loans_and_advances']['entries'],
                        'total': balance_sheet['assets']['current_assets']['loans_and_advances']['total']
                    },
                    'other_assets': {
                        'title': 'Other Assets',
                        'entries': balance_sheet['assets']['current_assets']['other_current_assets']['entries'],
                        'total': balance_sheet['assets']['current_assets']['other_current_assets']['total']
                    }
                },
                'total': (
                    balance_sheet['assets']['current_assets']['loans_and_advances']['total'] +
                    balance_sheet['assets']['current_assets']['other_current_assets']['total']
                )
            }
        }

        # Debug logging
        print("\nDebug - Generated Notes Structure:")
        for note_key, note_data in notes.items():
            print(f"\n{note_data['title']}:")
            if 'sub_categories' in note_data:
                for sub_cat_key, sub_cat_data in note_data['sub_categories'].items():
                    print(f"  {sub_cat_data['title']}: {sub_cat_data['total']}")
            print(f"Total: {note_data.get('total', 0)}")
        
        return notes
        
    except Exception as e:
        print(f"Error in generate_balance_sheet_notes: {str(e)}")
        print(f"Balance sheet structure: {balance_sheet}")
        raise    

@financials.route('/balance_sheet_notes/<client_id>')
@login_required
def view_balance_sheet_notes(client_id):
    try:
        print("\n=== Starting view_balance_sheet_notes ===")
        db = get_db()
        
        # 1. Get client info with logging
        print(f"Looking up client with ID: {client_id}")
        client = db.financial_clients.find_one({
            '_id': ObjectId(client_id),
            'user_id': session['user_id']
        })
        
        if not client:
            print("Client not found in database")
            flash('Client not found', 'error')
            return redirect(url_for('financials.financial_dashboard'))
        print("Client found:", client.get('legal_name'))
        
        # 2. Get trial balance with logging
        print("Looking up trial balance")
        trial_balance = db.trial_balances.find_one({
            'client_id': ObjectId(client_id),
            'user_id': session['user_id']
        })
        
        if not trial_balance:
            print("Trial balance not found")
            flash('Trial balance not found', 'error')
            return redirect(url_for('financials.view_client_details', client_id=client_id))
        print("Trial balance found")
        
        # 3. Handle date conversion
        try:
            if isinstance(client['end_date'], str):
                balance_sheet_date = datetime.strptime(client['end_date'], '%Y-%m-%d')
            else:
                balance_sheet_date = client['end_date']
            print(f"Balance sheet date: {balance_sheet_date}")
        except Exception as e:
            print(f"Date conversion error: {str(e)}")
            balance_sheet_date = datetime.now()

        # 4. Generate balance sheet with logging
        print("Generating balance sheet")
        try:
            balance_sheet = generate_balance_sheet(
                client_id, 
                trial_balance['data'],
                balance_sheet_date
            )
            print("Balance sheet generated successfully")
        except Exception as e:
            print(f"Balance sheet generation error: {str(e)}")
            print(f"Trial balance data: {trial_balance['data']}")
            raise

        # 5. Generate notes with logging
        print("Generating balance sheet notes")
        try:
            notes = generate_balance_sheet_notes(trial_balance['data'], balance_sheet)
            print("Notes generated successfully")
        except Exception as e:
            print(f"Notes generation error: {str(e)}")
            print(f"Balance sheet structure: {balance_sheet}")
            raise

        # 6. Render template with logging
        print("Rendering template")
        try:
            return render_template(
                'financials/balance_sheet_notes.html',
                client=client,
                notes=notes,
                balance_sheet_date=balance_sheet_date,
                format_amount=lambda x: format_amount(
                    x, 
                    client['settings']['rounding_amount'], 
                    int(client['settings'].get('decimal_rounding', 2))
                )
            )
        except Exception as e:
            print(f"Template rendering error: {str(e)}")
            raise
            
    except Exception as e:
        import traceback
        print("\n=== Error in view_balance_sheet_notes ===")
        print(f"Error type: {type(e)}")
        print(f"Error message: {str(e)}")
        print("Full traceback:")
        print(traceback.format_exc())
        flash('Error accessing balance sheet notes', 'error')
        return redirect(url_for('financials.financial_dashboard'))
    
def generate_profit_loss_notes(trial_balance_data):
    """Generate detailed notes for profit and loss statement with subschedule grouping"""
    try:
        notes = {
            # Note 12: Revenue from Operations
            'note_12': {
                'title': 'Note 12: Revenue from Operations',
                'entries': [],
                'total': 0
            },
            
            # Note 13: Other Income
            'note_13': {
                'title': 'Note 13: Other Income',
                'entries': [],
                'total': 0
            },
            
            # Note 14: Cost of Materials
            'note_14': {
                'title': 'Note 14: Cost of Materials Consumed',
                'entries': [],
                'total': 0
            },
            
            # Note 15: Employee Benefits
            'note_15': {
                'title': 'Note 15: Employee Benefits Expense',
                'sub_categories': {
                    'salaries_and_wages': {
                        'title': 'Salaries and Wages',
                        'entries': [],
                        'total': 0
                    },
                    'staff_welfare': {
                        'title': 'Staff Welfare Expenses',
                        'entries': [],
                        'total': 0
                    },
                    'bonus': {
                        'title': 'Bonus',
                        'entries': [],
                        'total': 0
                    }
                },
                'total': 0
            },
            
            # Note 16: Finance Costs
            'note_16': {
                'title': 'Note 16: Finance Costs',
                'sub_categories': {
                    'interest_expense': {
                        'title': 'Interest Expense',
                        'entries': [],
                        'total': 0
                    },
                    'bank_charges': {
                        'title': 'Bank Charges',
                        'entries': [],
                        'total': 0
                    },
                    'other_borrowing_costs': {
                        'title': 'Other Borrowing Costs',
                        'entries': [],
                        'total': 0
                    }
                },
                'total': 0
            },
            
            # Note 17: Other Expenses
            'note_17': {
                'title': 'Note 17: Other Expenses',
                'entries': [],
                'total': 0
            }
        }

        # Process trial balance entries
        for entry in trial_balance_data:
            try:
                debit = float(entry.get('Debit', 0) or 0)
                credit = float(entry.get('Credit', 0) or 0)
                mapping = str(entry.get('Mapping', '')).strip()
                account_name = str(entry.get('Account_Name', '')).strip()
                
                if not mapping.startswith('PL-'):
                    continue
                    
                # Get main category and section
                main_category = get_main_pl_category(mapping)
                section = get_pl_section(mapping)
                
                if not main_category or not section:
                    continue
                
                # Calculate amount based on section
                if section == 'income':
                    amount = credit - debit
                else:
                    amount = debit - credit
                    
                # Add entry to appropriate note
                if mapping in PL_MAPPING_STRUCTURE['revenue_from_operations']['subschedules']:
                    notes['note_12']['entries'].append({
                        'account': account_name,
                        'amount': amount,
                        'subschedule': mapping
                    })
                    notes['note_12']['total'] += amount
                    
                elif mapping in PL_MAPPING_STRUCTURE['other_income']['subschedules']:
                    notes['note_13']['entries'].append({
                        'account': account_name,
                        'amount': amount,
                        'subschedule': mapping
                    })
                    notes['note_13']['total'] += amount
                    
                elif mapping in PL_MAPPING_STRUCTURE['cost_of_materials']['subschedules']:
                    notes['note_14']['entries'].append({
                        'account': account_name,
                        'amount': amount,
                        'subschedule': mapping
                    })
                    notes['note_14']['total'] += amount
                    
                elif mapping in PL_MAPPING_STRUCTURE['employee_benefits']['subschedules']:
                    # Categorize employee benefits based on account name
                    if 'salary' in account_name.lower() or 'wage' in account_name.lower():
                        notes['note_15']['sub_categories']['salaries_and_wages']['entries'].append({
                            'account': account_name,
                            'amount': amount
                        })
                        notes['note_15']['sub_categories']['salaries_and_wages']['total'] += amount
                    elif 'welfare' in account_name.lower():
                        notes['note_15']['sub_categories']['staff_welfare']['entries'].append({
                            'account': account_name,
                            'amount': amount
                        })
                        notes['note_15']['sub_categories']['staff_welfare']['total'] += amount
                    elif 'bonus' in account_name.lower():
                        notes['note_15']['sub_categories']['bonus']['entries'].append({
                            'account': account_name,
                            'amount': amount
                        })
                        notes['note_15']['sub_categories']['bonus']['total'] += amount
                    notes['note_15']['total'] += amount
                    
                elif mapping in PL_MAPPING_STRUCTURE['finance_costs']['subschedules']:
                    # Categorize finance costs
                    if 'interest' in account_name.lower():
                        notes['note_16']['sub_categories']['interest_expense']['entries'].append({
                            'account': account_name,
                            'amount': amount
                        })
                        notes['note_16']['sub_categories']['interest_expense']['total'] += amount
                    elif 'bank charge' in account_name.lower():
                        notes['note_16']['sub_categories']['bank_charges']['entries'].append({
                            'account': account_name,
                            'amount': amount
                        })
                        notes['note_16']['sub_categories']['bank_charges']['total'] += amount
                    else:
                        notes['note_16']['sub_categories']['other_borrowing_costs']['entries'].append({
                            'account': account_name,
                            'amount': amount
                        })
                        notes['note_16']['sub_categories']['other_borrowing_costs']['total'] += amount
                    notes['note_16']['total'] += amount
                    
                elif mapping in PL_MAPPING_STRUCTURE['other_expenses']['subschedules']:
                    notes['note_17']['entries'].append({
                        'account': account_name,
                        'amount': amount,
                        'subschedule': mapping
                    })
                    notes['note_17']['total'] += amount

            except Exception as e:
                print(f"Error processing entry for notes: {str(e)}")
                continue

        # Sort entries within each note by amount descending
        for note in notes.values():
            if 'entries' in note:
                note['entries'].sort(key=lambda x: x.get('amount', 0), reverse=True)
            elif 'sub_categories' in note:
                for sub_cat in note['sub_categories'].values():
                    sub_cat['entries'].sort(key=lambda x: x.get('amount', 0), reverse=True)

        return notes
        
    except Exception as e:
        print(f"Error generating profit and loss notes: {str(e)}")
        raise
    
@financials.route('/profit_loss_notes/<client_id>')
@login_required
def view_profit_loss_notes(client_id):
    try:
        db = get_db()
        client = db.financial_clients.find_one({
            '_id': ObjectId(client_id),
            'user_id': session['user_id']
        })
        
        if not client:
            flash('Client not found', 'error')
            return redirect(url_for('financials.financial_dashboard'))
        
        trial_balance = db.trial_balances.find_one({
            'client_id': ObjectId(client_id),
            'user_id': session['user_id']
        })
        
        if not trial_balance:
            flash('Trial balance not found', 'error')
            return redirect(url_for('financials.view_client_details', client_id=client_id))

        notes = generate_profit_loss_notes(trial_balance['data'])
        
        return render_template(
            'financials/profit_loss_notes.html',
            client=client,
            notes=notes,
            format_amount=lambda x: format_amount(
                x, 
                client['settings']['rounding_amount'], 
                int(client['settings'].get('decimal_rounding', 2))
            )
        )
        
    except Exception as e:
        print(f"Error viewing profit and loss notes: {str(e)}")
        flash('Error accessing profit and loss notes', 'error')
        return redirect(url_for('financials.financial_dashboard'))
    
@financials.route('/export_cash_flow_excel/<client_id>')
@login_required
def export_cash_flow_excel(client_id):
    try:
        db = get_db()
        client = db.financial_clients.find_one({
            '_id': ObjectId(client_id),
            'user_id': session['user_id']
        })
        
        if not client:
            flash('Client not found', 'error')
            return redirect(url_for('financials.financial_dashboard'))

        # Get balance sheet dates
        balance_sheet_date = datetime.strptime(client['end_date'], '%Y-%m-%d') if isinstance(client['end_date'], str) else client['end_date']
        previous_year_date = balance_sheet_date - timedelta(days=365)
        
        # Get trial balances
        current_trial_balance = db.trial_balances.find_one({
            'client_id': ObjectId(client_id),
            'user_id': session['user_id']
        })
        
        previous_trial_balance = db.trial_balances.find_one({
            'client_id': ObjectId(client_id),
            'user_id': session['user_id'],
            'period.end_date': previous_year_date.strftime('%Y-%m-%d')
        })

        # Generate required data
        current_balance_sheet = generate_balance_sheet(
            client_id, 
            current_trial_balance['data'],
            balance_sheet_date
        )
        
        previous_balance_sheet = generate_balance_sheet(
            client_id,
            previous_trial_balance['data'] if previous_trial_balance else [],
            previous_year_date
        )
        
        profit_loss_data = generate_profit_loss_statement(current_trial_balance['data'])
        
        cash_flow_data = generate_cash_flow_statement(
            current_balance_sheet,
            previous_balance_sheet,
            profit_loss_data
        )

        # Create Excel workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Cash Flow Statement"
        
        # Add headers
        ws['A1'] = client['legal_name']
        ws['A2'] = 'Statement of Cash Flows'
        ws['A3'] = f'Year ended March 31, {client["fiscal_year"]}'
        ws['A4'] = f'All amounts in INR {client["settings"]["rounding_amount"]}, unless otherwise stated'
        
        # Format headers
        for i in range(1, 5):
            cell = ws[f'A{i}']
            cell.font = Font(bold=True)
            if i == 1:
                cell.font.size = 14
            ws.merge_cells(f'A{i}:D{i}')
            cell.alignment = Alignment(horizontal='center')

        # Add column headers
        headers = ['Particulars', f'Year ended March 31, {client["fiscal_year"]}', 
                  f'Year ended March 31, {int(client["fiscal_year"]) - 1}']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=6, column=col)
            cell.value = header
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="E6E6E6", end_color="E6E6E6", fill_type="solid")

        current_row = 7

        # A. Operating Activities
        ws.cell(row=current_row, column=1, value="A. CASH FLOW FROM OPERATING ACTIVITIES").font = Font(bold=True)
        current_row += 1

        operating_items = [
            ('Profit before tax', cash_flow_data['operating_activities']['profit_before_tax']),
            ('Increase / (Decrease) in Trade Payables', cash_flow_data['operating_activities']['changes']['trade_payables']),
            ('Increase / (Decrease) in Other liabilities', cash_flow_data['operating_activities']['changes']['other_liabilities']),
            ('Decrease / (Increase) in Inventories', cash_flow_data['operating_activities']['changes']['inventories']),
            ('Decrease / (Increase) in Trade Receivables', cash_flow_data['operating_activities']['changes']['trade_receivables']),
            ('Decrease / (Increase) in loans and advances', cash_flow_data['operating_activities']['changes']['loans_advances']),
            ('Decrease / (Increase) in Other assets', cash_flow_data['operating_activities']['changes']['other_assets']),
            ('Cash generated from / (used in) Operations', cash_flow_data['operating_activities']['cash_generated']),
            ('Net Cash generated from / (used in) Operating Activities', cash_flow_data['operating_activities']['net_cash'])
        ]

        for item in operating_items:
            ws.cell(row=current_row, column=1, value=item[0])
            ws.cell(row=current_row, column=2, value=item[1]).number_format = '#,##0.00'
            current_row += 1

        current_row += 1

        # In the export_cash_flow_excel function, add this section after operating activities
        # B. Investing Activities
        current_row += 1
        ws.cell(row=current_row, column=1, value="B. CASH FLOW FROM INVESTING ACTIVITIES").font = Font(bold=True)
        current_row += 1

        investing_items = []
        if cash_flow_data['investing_activities']['fixed_assets_purchase'] > 0:
            investing_items.append(('Purchase of Fixed Assets', -cash_flow_data['investing_activities']['fixed_assets_purchase']))
        if cash_flow_data['investing_activities']['fixed_assets_sale'] > 0:
            investing_items.append(('Sale of Fixed Assets', cash_flow_data['investing_activities']['fixed_assets_sale']))
        if cash_flow_data['investing_activities']['investments_made'] > 0:
            investing_items.append(('Purchase of Investments', -cash_flow_data['investing_activities']['investments_made']))
        if cash_flow_data['investing_activities']['investments_sale'] > 0:
            investing_items.append(('Sale of Investments', cash_flow_data['investing_activities']['investments_sale']))
        investing_items.append(('Net Cash generated from / (used in) Investing Activities', 
                            cash_flow_data['investing_activities']['net_cash']))

        for item in investing_items:
            ws.cell(row=current_row, column=1, value=item[0])
            ws.cell(row=current_row, column=2, value=item[1]).number_format = '#,##0.00'
            current_row += 1

        # C. Financing Activities
        ws.cell(row=current_row, column=1, value="C. CASH FLOW FROM FINANCING ACTIVITIES").font = Font(bold=True)
        current_row += 1

        financing_items = [
            ('Proceeds from issue of Share capital', cash_flow_data['financing_activities']['share_capital']),
            ('Proceeds from Long-Term Borrowings', cash_flow_data['financing_activities']['long_term_borrowings']),
            ('Net Cash generated from / (used in) Financing Activities', cash_flow_data['financing_activities']['net_cash'])
        ]

        for item in financing_items:
            ws.cell(row=current_row, column=1, value=item[0])
            ws.cell(row=current_row, column=2, value=item[1]).number_format = '#,##0.00'
            current_row += 1

        current_row += 1

        # Net Change in Cash
        final_items = [
            ('Net Increase / (Decrease) in Cash and Cash Equivalents', cash_flow_data['cash_position']['net_increase']),
            ('Cash and Cash Equivalents at the Beginning', cash_flow_data['cash_position']['beginning_balance']),
            ('Cash and Cash Equivalents at the End', cash_flow_data['cash_position']['ending_balance'])
        ]

        for item in final_items:
            ws.cell(row=current_row, column=1, value=item[0]).font = Font(bold=True)
            ws.cell(row=current_row, column=2, value=item[1]).number_format = '#,##0.00'
            current_row += 1

        # Adjust column widths
        ws.column_dimensions['A'].width = 60
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 20

        # Create response
        excel_file = BytesIO()
        wb.save(excel_file)
        excel_file.seek(0)

        return send_file(
            excel_file,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=f'Cash_Flow_{client["legal_name"]}_{client["fiscal_year"]}.xlsx'
        )

    except Exception as e:
        print(f"Error exporting cash flow to Excel: {str(e)}")
        flash(f'Error exporting to Excel: {str(e)}', 'error')
        return redirect(url_for('financials.generate_cash_flow', client_id=client_id))


# Helper function to safely format dates
def format_date(date_value):
    if not date_value:
        return "N/A"
    try:
        if isinstance(date_value, str):
            return datetime.strptime(date_value, '%Y-%m-%d').strftime('%B %d, %Y')
        elif isinstance(date_value, datetime):
            return date_value.strftime('%B %d, %Y')
        return str(date_value)
    except:
        return str(date_value)
@financials.route('/export_all_financials_complete/<client_id>')
@login_required
def export_all_financials_complete(client_id):
    try:
        db = get_db()
        client = db.financial_clients.find_one({'_id': ObjectId(client_id)})
        trial_balance = db.trial_balances.find_one({
            'client_id': ObjectId(client_id),
            'user_id': session['user_id']
        })
        
        if not trial_balance or not client:
            flash('Required data not found', 'error')
            return redirect(url_for('financials.view_client_details', client_id=client_id))

        # Initial date setup
        balance_sheet_date = datetime.strptime(client['end_date'], '%Y-%m-%d') if isinstance(client['end_date'], str) else client['end_date']
        previous_year_date = balance_sheet_date - timedelta(days=365)

        # Get previous year's trial balance
        previous_trial_balance = db.trial_balances.find_one({
            'client_id': ObjectId(client_id),
            'user_id': session['user_id'],
            'period.end_date': previous_year_date.strftime('%Y-%m-%d')
        })

        # Generate all required financial data
        balance_sheet = generate_balance_sheet(client_id, trial_balance['data'], balance_sheet_date)
        profit_loss_data = generate_profit_loss_statement(trial_balance['data'])
        
        # Generate previous year balance sheet
        previous_balance_sheet = generate_balance_sheet(
            client_id,
            previous_trial_balance['data'] if previous_trial_balance else [],
            previous_year_date
        ) if previous_trial_balance else get_empty_balance_sheet()

        # Generate cash flow data
        cash_flow_data = generate_cash_flow_statement(
            current_balance_sheet=balance_sheet,
            previous_balance_sheet=previous_balance_sheet,
            profit_loss_data=profit_loss_data
        )

        # Generate balance sheet notes data
        balance_sheet_notes = generate_balance_sheet_notes(trial_balance['data'], balance_sheet)

        # Generate profit and loss notes data
        profit_loss_notes = generate_profit_loss_notes(trial_balance['data'])

        wb = Workbook()
        
        # Styles
        title_style = NamedStyle(
            name='title_style',
            font=Font(bold=True, size=14),
            alignment=Alignment(horizontal='center'),
            border=Border(bottom=Side(style='thin'))
        )
        
        header_style = NamedStyle(
            name='header_style',
            font=Font(bold=True),
            fill=PatternFill(start_color='E0E0E0', end_color='E0E0E0', fill_type='solid'),
            alignment=Alignment(horizontal='left', vertical='center'),
            border=Border(bottom=Side(style='thin'))
        )
        
        subheader_style = NamedStyle(
            name='subheader_style',
            font=Font(bold=True),
            alignment=Alignment(horizontal='left'),
            border=Border(bottom=Side(style='thin'))
        )
        
        number_style = NamedStyle(
            name='number_style',
            number_format='#,##0.00',
            alignment=Alignment(horizontal='right'),
            border=Border(bottom=Side(style='thin'))
        )

        # Balance Sheet
        bs = wb.active
        bs.title = "Balance Sheet"
        # Set print options
        bs.page_setup.orientation = bs.ORIENTATION_PORTRAIT
        bs.page_setup.paperSize = bs.PAPERSIZE_A4
        bs.page_setup.fitToPage = True
        bs.page_setup.fitToHeight = 1
        bs.page_setup.fitToWidth = 1
        
        # Set margins (in inches)
        bs.page_margins.left = 0.5
        bs.page_margins.right = 0.5
        bs.page_margins.top = 0.5
        bs.page_margins.bottom = 0.5
        bs.page_margins.header = 0.3
        bs.page_margins.footer = 0.3

        # Adjust column widths precisely
        bs.column_dimensions['A'].width = 45  # Particulars
        bs.column_dimensions['B'].width = 8   # Note
        bs.column_dimensions['C'].width = 15  # Current year
        bs.column_dimensions['D'].width = 15  # Previous year

        # Adjust row heights
        default_row_height = 15  # in points
        for row in range(1, bs.max_row + 1):
            bs.row_dimensions[row].height = default_row_height

        # Adjust font size
        base_font_size = 9
        title_font_size = 11
        
        title_style = NamedStyle(
            name='title_style',
            font=Font(bold=True, size=title_font_size),
            alignment=Alignment(horizontal='center', vertical='center')
        )
        
        header_style = NamedStyle(
            name='header_style',
            font=Font(bold=True, size=base_font_size),
            fill=PatternFill(start_color='E0E0E0', end_color='E0E0E0', fill_type='solid'),
            alignment=Alignment(horizontal='left', vertical='center')
        )
        
        subheader_style = NamedStyle(
            name='subheader_style',
            font=Font(bold=True, size=base_font_size),
            alignment=Alignment(horizontal='left', vertical='center')
        )
        
        number_style = NamedStyle(
            name='number_style',
            font=Font(size=base_font_size),
            number_format='#,##0.00',
            alignment=Alignment(horizontal='right', vertical='center')
        )

        # Compact the spacing between sections
        # Reduce the number of blank rows between sections to 1 instead of 2
        # Rest of your existing balance sheet code remains the same, but wherever you have:
        # current_row += 2
        # Change it to:
        # current_row += 1

        # Add print title rows (headers that repeat on each page)


        # Add page setup options for better printing
        bs.page_setup.horizontalCentered = True
        bs.page_setup.verticalCentered = True

        # Add print grid lines
        bs.print_options.gridLines = True

        # Optimize row heights based on content
        for row in range(1, bs.max_row + 1):
            bs.row_dimensions[row].height = 15  # Set a consistent height

        # Set zoom level for better view
        bs.sheet_view.zoomScale = 100
        
        
        # Headers
        bs['A1'] = f"{client['legal_name']} - Balance Sheet"
        bs['A1'].style = title_style
        bs.merge_cells('A1:D1')
        
        bs['A2'] = f"As at {balance_sheet_date.strftime('%B %d, %Y')}"
        bs['A2'].style = subheader_style
        bs.merge_cells('A2:D2')
        
        bs['A3'] = f"All amounts in {client['settings']['rounding_amount']}"
        bs.merge_cells('A3:D3')

        current_row = 5
        headers = ["Particulars", "Note", f"As at {balance_sheet_date.strftime('%B %d, %Y')}", 
                  f"As at {previous_year_date.strftime('%B %d, %Y')}"]
        for col, header in enumerate(headers, 1):
            bs.cell(row=current_row, column=col, value=header).style = header_style

        current_row += 1

        # (1) Equity
        bs[f'A{current_row}'] = "(1) Equity"
        bs[f'A{current_row}'].style = subheader_style
        current_row += 1

        # Share Capital
        bs[f'A{current_row}'] = "Share Capital"
        bs[f'B{current_row}'] = "3"
        bs[f'C{current_row}'] = get_bs_value(balance_sheet, 'liabilities.equity.equity_share_capital')
        bs[f'D{current_row}'] = get_bs_value(previous_balance_sheet, 'liabilities.equity.equity_share_capital')
        apply_number_style(bs, current_row, ['C', 'D'])
        current_row += 1

        # Securities Premium
        bs[f'A{current_row}'] = "Securities Premium"
        bs[f'B{current_row}'] = "3"
        bs[f'C{current_row}'] = get_bs_value(balance_sheet, 'liabilities.equity.securities_premium')
        bs[f'D{current_row}'] = get_bs_value(previous_balance_sheet, 'liabilities.equity.securities_premium')
        apply_number_style(bs, current_row, ['C', 'D'])
        current_row += 1

        # General Reserve
        bs[f'A{current_row}'] = "General Reserve"
        bs[f'B{current_row}'] = "3"
        bs[f'C{current_row}'] = get_bs_value(balance_sheet, 'liabilities.equity.general_reserve')
        bs[f'D{current_row}'] = get_bs_value(previous_balance_sheet, 'liabilities.equity.general_reserve')
        apply_number_style(bs, current_row, ['C', 'D'])
        current_row += 1

        # Capital Reserve
        bs[f'A{current_row}'] = "Capital Reserve"
        bs[f'B{current_row}'] = "3"
        bs[f'C{current_row}'] = get_bs_value(balance_sheet, 'liabilities.equity.capital_reserve')
        bs[f'D{current_row}'] = get_bs_value(previous_balance_sheet, 'liabilities.equity.capital_reserve')
        apply_number_style(bs, current_row, ['C', 'D'])
        current_row += 1

        # Retained Earnings
        bs[f'A{current_row}'] = "Opening Retained Earnings"
        bs[f'B{current_row}'] = "3"
        bs[f'C{current_row}'] = get_bs_value(balance_sheet, 'liabilities.equity.retained_earnings.opening_balance')
        bs[f'D{current_row}'] = get_bs_value(previous_balance_sheet, 'liabilities.equity.retained_earnings.opening_balance')
        apply_number_style(bs, current_row, ['C', 'D'])
        current_row += 1

        # Add: Profit for the year
        bs[f'A{current_row}'] = "Add: Profit for the year"
        bs[f'B{current_row}'] = "3"
        bs[f'C{current_row}'] = get_bs_value(balance_sheet, 'liabilities.equity.retained_earnings.current_year_profit')
        bs[f'D{current_row}'] = get_bs_value(previous_balance_sheet, 'liabilities.equity.retained_earnings.current_year_profit')
        apply_number_style(bs, current_row, ['C', 'D'])
        current_row += 1

        # Total Retained Earnings
        bs[f'A{current_row}'] = "Total Retained Earnings"
        bs[f'C{current_row}'] = get_bs_value(balance_sheet, 'liabilities.equity.retained_earnings.total')
        bs[f'D{current_row}'] = get_bs_value(previous_balance_sheet, 'liabilities.equity.retained_earnings.total')
        apply_number_style(bs, current_row, ['C', 'D'], bold=True)
        current_row += 1

        # Total Equity
        bs[f'A{current_row}'] = "Total Equity"
        bs[f'C{current_row}'] = get_bs_value(balance_sheet, 'liabilities.equity.total')
        bs[f'D{current_row}'] = get_bs_value(previous_balance_sheet, 'liabilities.equity.total')
        apply_number_style(bs, current_row, ['C', 'D'], bold=True)
        current_row += 2

        # (2) Non-Current Liabilities
        bs[f'A{current_row}'] = "(2) Non-Current Liabilities"
        bs[f'A{current_row}'].style = subheader_style
        current_row += 1

        # Secured Loans
        bs[f'A{current_row}'] = "Secured Loans"
        bs[f'B{current_row}'] = "4"
        bs[f'C{current_row}'] = get_bs_value(balance_sheet, 'liabilities.long_term_liabilities.secured_loans')
        bs[f'D{current_row}'] = get_bs_value(previous_balance_sheet, 'liabilities.long_term_liabilities.secured_loans')
        apply_number_style(bs, current_row, ['C', 'D'])
        current_row += 1

        # Unsecured Loans
        bs[f'A{current_row}'] = "Unsecured Loans"
        bs[f'B{current_row}'] = "4"
        bs[f'C{current_row}'] = get_bs_value(balance_sheet, 'liabilities.long_term_liabilities.unsecured_loans')
        bs[f'D{current_row}'] = get_bs_value(previous_balance_sheet, 'liabilities.long_term_liabilities.unsecured_loans')
        apply_number_style(bs, current_row, ['C', 'D'])
        current_row += 1

        # Total Non-Current Liabilities
        bs[f'A{current_row}'] = "Total Non-Current Liabilities"
        bs[f'C{current_row}'] = get_bs_value(balance_sheet, 'liabilities.long_term_liabilities.total')
        bs[f'D{current_row}'] = get_bs_value(previous_balance_sheet, 'liabilities.long_term_liabilities.total')
        apply_number_style(bs, current_row, ['C', 'D'], bold=True)
        current_row += 2

        # (3) Current Liabilities
        bs[f'A{current_row}'] = "(3) Current Liabilities"
        bs[f'A{current_row}'].style = subheader_style
        current_row += 1

        # Current Liabilities items
        current_liabilities = [
            ("Trade Payables - MSME", "5", "msme_payables"),
            ("Trade Payables - Others", "5", "other_payables"),
            ("Statutory Dues", "6", "statutory_dues"),
            ("Short Term Borrowings", "6", "short_term_borrowings"),
            ("Other Current Liabilities", "6", "other_current_liabilities")
        ]

        for label, note, key in current_liabilities:
            bs[f'A{current_row}'] = label
            bs[f'B{current_row}'] = note
            bs[f'C{current_row}'] = get_bs_value(balance_sheet, f'liabilities.current_liabilities.{key}')
            bs[f'D{current_row}'] = get_bs_value(previous_balance_sheet, f'liabilities.current_liabilities.{key}')
            apply_number_style(bs, current_row, ['C', 'D'])
            current_row += 1

        # Total Current Liabilities
        bs[f'A{current_row}'] = "Total Current Liabilities"
        bs[f'C{current_row}'] = get_bs_value(balance_sheet, 'liabilities.current_liabilities.total')
        bs[f'D{current_row}'] = get_bs_value(previous_balance_sheet, 'liabilities.current_liabilities.total')
        apply_number_style(bs, current_row, ['C', 'D'], bold=True)
        current_row += 1

        # TOTAL EQUITY AND LIABILITIES
        bs[f'A{current_row}'] = "TOTAL EQUITY AND LIABILITIES"
        bs[f'C{current_row}'] = get_bs_value(balance_sheet, 'total_liabilities')
        bs[f'D{current_row}'] = get_bs_value(previous_balance_sheet, 'total_liabilities')
        apply_number_style(bs, current_row, ['C', 'D'], bold=True)
        current_row += 2

        # II. ASSETS
        bs[f'A{current_row}'] = "II. ASSETS"
        bs[f'A{current_row}'].style = header_style
        bs.merge_cells(f'A{current_row}:D{current_row}')
        current_row += 1

        # (1) Non-Current Assets
        bs[f'A{current_row}'] = "(1) Non-Current Assets"
        bs[f'A{current_row}'].style = subheader_style
        current_row += 1

        # Fixed Assets
        fixed_assets = [
            ("Tangible Assets", "7", "tangible_assets"),
            ("Intangible Assets", "7", "intangible_assets")
        ]

        for label, note, key in fixed_assets:
            bs[f'A{current_row}'] = label
            bs[f'B{current_row}'] = note
            bs[f'C{current_row}'] = get_bs_value(balance_sheet, f'assets.fixed_assets.{key}')
            bs[f'D{current_row}'] = get_bs_value(previous_balance_sheet, f'assets.fixed_assets.{key}')
            apply_number_style(bs, current_row, ['C', 'D'])
            current_row += 1

        # Total Fixed Assets
        bs[f'A{current_row}'] = "Total Fixed Assets"
        bs[f'C{current_row}'] = get_bs_value(balance_sheet, 'assets.fixed_assets.total')
        bs[f'D{current_row}'] = get_bs_value(previous_balance_sheet, 'assets.fixed_assets.total')
        apply_number_style(bs, current_row, ['C', 'D'], bold=True)
        current_row += 1

        # Long Term Investments
        bs[f'A{current_row}'] = "Long Term Investments"
        bs[f'B{current_row}'] = "8"
        bs[f'C{current_row}'] = get_bs_value(balance_sheet, 'assets.investments.long_term_investments')
        bs[f'D{current_row}'] = get_bs_value(previous_balance_sheet, 'assets.investments.long_term_investments')
        apply_number_style(bs, current_row, ['C', 'D'])
        current_row += 2

        # (2) Current Assets
        bs[f'A{current_row}'] = "(2) Current Assets"
        bs[f'A{current_row}'].style = subheader_style
        current_row += 1

        # Debug print to check values
        print("Debug: Balance Sheet Structure")
        print(f"Short Term Investments: {get_bs_value(balance_sheet, 'assets.investments.short_term_investments.total')}")
        print(f"Inventories: {get_bs_value(balance_sheet, 'assets.current_assets.inventories.total')}")
        print(f"Trade Receivables: {get_bs_value(balance_sheet, 'assets.current_assets.trade_receivables.total')}")

        # Current Assets items with corrected paths and structure
        current_assets = [
            ("Short Term Investments", "8", "assets.investments.short_term_investments.total"),
            ("Inventories", "9", "assets.current_assets.inventories.total"),
            ("Trade Receivables", "10", "assets.current_assets.trade_receivables.total"),
            ("Cash and Cash Equivalents", "11", "assets.current_assets.cash_and_equivalents.total"),
            ("Loans and Advances", "12", "assets.current_assets.loans_and_advances.total"),
            ("Other Current Assets", "12", "assets.current_assets.other_current_assets.total")
        ]

        for label, note, key in current_assets:
            try:
                value = get_bs_value(balance_sheet, key)
                prev_value = get_bs_value(previous_balance_sheet, key)
                
                print(f"Debug: {label} - Current: {value}, Previous: {prev_value}")  # Debug print
                
                bs[f'A{current_row}'] = label
                bs[f'B{current_row}'] = note
                bs[f'C{current_row}'] = value
                bs[f'D{current_row}'] = prev_value
                apply_number_style(bs, current_row, ['C', 'D'])
                current_row += 1
            except Exception as e:
                print(f"Error processing {label}: {str(e)}")

        # Add Total Current Assets
        bs[f'A{current_row}'] = "Total Current Assets"
        bs[f'A{current_row}'].font = Font(bold=True)
        bs[f'C{current_row}'] = get_bs_value(balance_sheet, 'assets.current_assets.total')
        bs[f'D{current_row}'] = get_bs_value(previous_balance_sheet, 'assets.current_assets.total')
        apply_number_style(bs, current_row, ['C', 'D'], bold=True)
        current_row += 1

        # Add TOTAL ASSETS
        bs[f'A{current_row}'] = "TOTAL ASSETS"
        bs[f'A{current_row}'].font = Font(bold=True)
        bs[f'C{current_row}'] = get_bs_value(balance_sheet, 'total_assets')
        bs[f'D{current_row}'] = get_bs_value(previous_balance_sheet, 'total_assets')
        apply_number_style(bs, current_row, ['C', 'D'], bold=True)
        current_row += 2

# After TOTAL ASSETS section, add the signatures
        current_row += 3

        # Signature section
        # Header
        bs[f'A{current_row}'] = "As per our report of even date attached"
        bs[f'A{current_row}'].style = header_style
        bs.merge_cells(f'A{current_row}:D{current_row}')
        current_row += 2

        # Store starting row for signature blocks
        signature_start_row = current_row

        # Left Side - Auditor Details
        auditor_col = 'A'
        bs[f'{auditor_col}{current_row}'] = f"For {client.get('audit_firm', {}).get('name', 'N/A')}"
        bs[f'{auditor_col}{current_row}'].font = Font(bold=True)
        current_row += 1

        bs[f'{auditor_col}{current_row}'] = "Chartered Accountants"
        current_row += 1

        bs[f'{auditor_col}{current_row}'] = f"Firm Registration No: {client.get('audit_firm', {}).get('frn', 'N/A')}"
        current_row += 2  # Extra space before partner details

        bs[f'{auditor_col}{current_row}'] = client.get('audit_firm', {}).get('auditor_name', 'N/A')
        bs[f'{auditor_col}{current_row}'].font = Font(bold=True)
        current_row += 1

        bs[f'{auditor_col}{current_row}'] = "Partner"
        current_row += 1

        bs[f'{auditor_col}{current_row}'] = f"Membership No: {client.get('audit_firm', {}).get('membership_no', 'N/A')}"
        current_row += 1

        bs[f'{auditor_col}{current_row}'] = f"UDIN: {client.get('audit_firm', {}).get('udin', 'N/A')}"
        current_row += 1

        # Right Side - Client Details
        client_col = 'D'
        right_start = signature_start_row
        
        bs[f'{client_col}{right_start}'] = f"For {client.get('legal_name', 'N/A')}"
        bs[f'{client_col}{right_start}'].font = Font(bold=True)
        
        bs[f'{client_col}{right_start + 4}'] = client.get('signatory', {}).get('name', 'N/A')
        bs[f'{client_col}{right_start + 4}'].font = Font(bold=True)
        
        bs[f'{client_col}{right_start + 5}'] = client.get('signatory', {}).get('designation', 'N/A')
        
        bs[f'{client_col}{right_start + 6}'] = f"PAN: {client.get('signatory', {}).get('pan', 'N/A')}"

        # Place and Date
        current_row += 2
        # Left side place and date
        bs[f'{auditor_col}{current_row}'] = f"Place: {client.get('audit_firm', {}).get('place', 'N/A')}"
        current_row += 1
        bs[f'{auditor_col}{current_row}'] = f"Date: {format_date(client.get('audit_firm', {}).get('date'))}"

        # Right side place and date
        bs[f'{client_col}{current_row-1}'] = f"Place: {client.get('signatory', {}).get('place', 'N/A')}"
        bs[f'{client_col}{current_row}'] = f"Date: {format_date(client.get('signatory', {}).get('date'))}"

        # Set column widths for better alignment
        bs.column_dimensions['A'].width = 35  # Left signature block
        bs.column_dimensions['B'].width = 15  # Spacing
        bs.column_dimensions['C'].width = 15  # Spacing
        bs.column_dimensions['D'].width = 35  # Right signature block

        # Calculate totals before verification
        total_equity_liabilities = get_bs_value(balance_sheet, 'total_liabilities')
        total_assets = get_bs_value(balance_sheet, 'total_assets')

        # Verify balance sheet equation
        if not math.isclose(total_equity_liabilities, total_assets, rel_tol=1e-9):
            print(f"Warning: Balance sheet not balanced! Total Equity & Liabilities: {total_equity_liabilities}, Total Assets: {total_assets}")
# Add this after all content generation (after signature section)
        
        # Set print titles and area
        bs.print_title_rows = '1:5'  # Repeat first 5 rows on each page
        bs.print_area = f'A1:D{current_row}'

        # Add page setup options for better printing
        bs.page_setup.horizontalCentered = True
        bs.page_setup.verticalCentered = True
        
        # Adjust zoom and grid lines
        bs.sheet_view.zoomScale = 100
        bs.print_options.gridLines = True

        # Final column width adjustments
        bs.column_dimensions['A'].width = 45  # Particulars
        bs.column_dimensions['B'].width = 8   # Note
        bs.column_dimensions['C'].width = 15  # Current year
        bs.column_dimensions['D'].width = 15  # Previous year

        # Calculate totals before verification
        total_equity_liabilities = get_bs_value(balance_sheet, 'total_liabilities')
        total_assets = get_bs_value(balance_sheet, 'total_assets')

        # Verify balance sheet equation
        if not math.isclose(total_equity_liabilities, total_assets, rel_tol=1e-9):
            print(f"Warning: Balance sheet not balanced! Total Equity & Liabilities: {total_equity_liabilities}, Total Assets: {total_assets}")

        # Create response
        excel_file = BytesIO()
        wb.save(excel_file)
        excel_file.seek(0)
# After your existing Balance Sheet code and before the response generation:

        # Create Profit and Loss Sheet
        pl = wb.create_sheet("Profit and Loss")

        # Headers
        pl['A1'] = f"{client['legal_name']} - Statement of Profit and Loss"
        pl['A1'].style = title_style
        pl.merge_cells('A1:D1')
        
        pl['A2'] = f"For the year ended {balance_sheet_date.strftime('%B %d, %Y')}"
        pl['A2'].style = subheader_style
        pl.merge_cells('A2:D2')
        
        pl['A3'] = f"All amounts in {client['settings']['rounding_amount']}"
        pl.merge_cells('A3:D3')

        pl_row = 5
        # Column headers
        headers = ["Particulars", "Note", f"Year ended {balance_sheet_date.strftime('%B %d, %Y')}", 
                  f"Year ended {previous_year_date.strftime('%B %d, %Y')}"]
        for col, header in enumerate(headers, 1):
            pl.cell(row=pl_row, column=col, value=header).style = header_style
        pl_row += 1

        # I. INCOME
        pl[f'A{pl_row}'] = "I. INCOME"
        pl[f'A{pl_row}'].style = header_style
        pl_row += 1

        # Revenue from operations
        pl[f'A{pl_row}'] = "(a) Revenue from operations"
        pl[f'B{pl_row}'] = "12"
        pl[f'C{pl_row}'] = profit_loss_data['income']['revenue_from_operations']
        pl[f'C{pl_row}'].style = number_style
        pl_row += 1

        # Other Income
        pl[f'A{pl_row}'] = "(b) Other Income"
        pl[f'B{pl_row}'] = "13"
        pl[f'C{pl_row}'] = profit_loss_data['income']['other_income']
        pl[f'C{pl_row}'].style = number_style
        pl_row += 1

        # Total Income
        pl[f'A{pl_row}'] = "Total Income"
        total_income = profit_loss_data['income']['total_income']
        pl[f'C{pl_row}'] = total_income
        pl[f'C{pl_row}'].style = number_style
        pl_row += 2

        # II. EXPENSES
        pl[f'A{pl_row}'] = "II. EXPENSES"
        pl[f'A{pl_row}'].style = header_style
        pl_row += 1

        # Cost of materials
        pl[f'A{pl_row}'] = "(a) Cost of materials consumed"
        pl[f'B{pl_row}'] = "14"
        pl[f'C{pl_row}'] = profit_loss_data['expenses']['cost_of_materials']
        pl[f'C{pl_row}'].style = number_style
        pl_row += 1

        # Employee benefits
        pl[f'A{pl_row}'] = "(b) Employee benefits expense"
        pl[f'B{pl_row}'] = "15"
        pl[f'C{pl_row}'] = profit_loss_data['expenses']['employee_benefits']
        pl[f'C{pl_row}'].style = number_style
        pl_row += 1

        # Finance costs
        pl[f'A{pl_row}'] = "(c) Finance costs"
        pl[f'B{pl_row}'] = "16"
        pl[f'C{pl_row}'] = profit_loss_data['expenses']['finance_costs']
        pl[f'C{pl_row}'].style = number_style
        pl_row += 1

        # Other expenses
        pl[f'A{pl_row}'] = "(d) Other expenses"
        pl[f'B{pl_row}'] = "17"
        pl[f'C{pl_row}'] = profit_loss_data['expenses']['other_expenses']
        pl[f'C{pl_row}'].style = number_style
        pl_row += 1

        # Total Expenses
        pl[f'A{pl_row}'] = "Total Expenses"
        total_expenses = profit_loss_data['expenses']['total_expenses']
        pl[f'C{pl_row}'] = total_expenses
        pl[f'C{pl_row}'].style = number_style
        pl_row += 2

        # III. Profit before tax
        pl[f'A{pl_row}'] = "III. Profit before tax (I - II)"
        pl[f'C{pl_row}'] = profit_loss_data['profit_before_tax']
        pl[f'C{pl_row}'].style = number_style
        pl_row += 2

        # IV. Tax Expenses
        pl[f'A{pl_row}'] = "IV. Tax Expenses"
        pl[f'A{pl_row}'].style = header_style
        pl_row += 1

        # Current tax
        pl[f'A{pl_row}'] = "(a) Current tax"
        pl[f'C{pl_row}'] = profit_loss_data['tax_expenses']['current_tax']
        pl[f'C{pl_row}'].style = number_style
        pl_row += 1

        # Total Tax Expenses
        pl[f'A{pl_row}'] = "Total Tax Expenses"
        pl[f'C{pl_row}'] = profit_loss_data['tax_expenses']['total_tax']
        pl[f'C{pl_row}'].style = number_style
        pl_row += 2

        # V. Profit for the year
        pl[f'A{pl_row}'] = "V. Profit for the year (III - IV)"
        pl[f'C{pl_row}'] = profit_loss_data['profit_for_year']
        pl[f'C{pl_row}'].style = number_style
        pl_row += 3

        # Add signatures section
        # Add signature section similar to Balance Sheet
        pl[f'A{pl_row}'] = "As per our report of even date attached"
        pl[f'A{pl_row}'].style = header_style
        pl.merge_cells(f'A{pl_row}:D{pl_row}')
        pl_row += 2

        # Left Side - Auditor Details
        pl[f'A{pl_row}'] = f"For {client.get('audit_firm', {}).get('name', 'N/A')}"
        pl[f'A{pl_row}'].font = Font(bold=True)
        pl_row += 1

        pl[f'A{pl_row}'] = "Chartered Accountants"
        pl_row += 1

        pl[f'A{pl_row}'] = f"Firm Registration No: {client.get('audit_firm', {}).get('frn', 'N/A')}"
        pl_row += 2

        # Partner Details
        pl[f'A{pl_row}'] = client.get('audit_firm', {}).get('auditor_name', 'N/A')
        pl[f'A{pl_row}'].font = Font(bold=True)
        pl_row += 1

        pl[f'A{pl_row}'] = "Partner"
        pl_row += 1

        pl[f'A{pl_row}'] = f"Membership No: {client.get('audit_firm', {}).get('membership_no', 'N/A')}"
        pl_row += 1

        # Right Side - Client Details
        signature_start = pl_row - 4
        pl[f'D{signature_start}'] = f"For {client.get('legal_name', 'N/A')}"
        pl[f'D{signature_start}'].font = Font(bold=True)

        pl[f'D{signature_start + 2}'] = client.get('signatory', {}).get('name', 'N/A')
        pl[f'D{signature_start + 2}'].font = Font(bold=True)
        pl[f'D{signature_start + 3}'] = client.get('signatory', {}).get('designation', 'N/A')

        # Place and Date
        pl_row += 2
        pl[f'A{pl_row}'] = f"Place: {client.get('signatory', {}).get('place', 'N/A')}"
        pl[f'D{pl_row}'] = f"Place: {client.get('audit_firm', {}).get('place', 'N/A')}"
        pl_row += 1
        pl[f'A{pl_row}'] = f"Date: {format_date(client.get('signatory', {}).get('date'))}"
        pl[f'D{pl_row}'] = f"Date: {format_date(client.get('audit_firm', {}).get('date'))}"

        # Set column widths
        pl.column_dimensions['A'].width = 45
        pl.column_dimensions['B'].width = 8
        pl.column_dimensions['C'].width = 15
        pl.column_dimensions['D'].width = 15

        # Set print area and other formatting
        pl.print_title_rows = '1:5'
        pl.print_area = f'A1:D{pl_row}'
        pl.page_setup.orientation = pl.ORIENTATION_PORTRAIT
        pl.page_setup.paperSize = pl.PAPERSIZE_A4
        pl.page_setup.fitToPage = True
        pl.page_setup.fitToHeight = 1
        pl.page_setup.fitToWidth = 1
        pl.page_setup.horizontalCentered = True
        pl.page_setup.verticalCentered = True
        pl.print_options.gridLines = True

        # Generate cash flow data
        cash_flow_data = generate_cash_flow_statement(
            balance_sheet,  # current year balance sheet
            previous_balance_sheet,
            profit_loss_data
        )

        # After the Profit & Loss sheet generation, add Cash Flow Statement
        cf = wb.create_sheet("Cash Flow Statement")

        # Headers
        cf['A1'] = f"{client['legal_name']} - Statement of Cash Flows"
        cf['A1'].style = title_style
        cf.merge_cells('A1:D1')

        cf['A2'] = f"For the year ended {balance_sheet_date.strftime('%B %d, %Y')}"
        cf['A2'].style = subheader_style
        cf.merge_cells('A2:D2')

        cf['A3'] = f"All amounts in {client['settings']['rounding_amount']}"
        cf.merge_cells('A3:D3')

        cf_row = 5
        headers = ["Particulars", "Note", f"Year ended {balance_sheet_date.strftime('%B %d, %Y')}", 
                f"Year ended {previous_year_date.strftime('%B %d, %Y')}"]
        for col, header in enumerate(headers, 1):
            cf.cell(row=cf_row, column=col, value=header).style = header_style
        cf_row += 1

        # A. Cash Flow from Operating Activities
        cf[f'A{cf_row}'] = "A. CASH FLOW FROM OPERATING ACTIVITIES"
        cf[f'A{cf_row}'].style = header_style
        cf_row += 1

        # Operating Activities Items
        # Operating Activities Items
        operating_items = [
            ('Profit Before Tax', cash_flow_data['operating_activities']['profit_before_tax']),
            ('Adjustments for:', None),
            # Change working_capital references instead of changes
            ('Changes in Trade Payables', cash_flow_data['operating_activities']['working_capital']['trade_payables']),
            ('Changes in Other Liabilities', cash_flow_data['operating_activities']['working_capital']['other_current_liabilities']),
            ('Changes in Trade Receivables', cash_flow_data['operating_activities']['working_capital']['trade_receivables']),
            ('Changes in Inventories', cash_flow_data['operating_activities']['working_capital']['inventories']),
            ('Changes in Other Assets', cash_flow_data['operating_activities']['working_capital']['other_current_assets'])
        ]

        for item in operating_items:
            cf[f'A{cf_row}'] = item[0]
            if item[1] is not None:
                cf[f'C{cf_row}'] = item[1]
                cf[f'C{cf_row}'].style = number_style
            cf_row += 1
        # Operating Activities Total
        cf[f'A{cf_row}'] = "Net Cash from Operating Activities (A)"
        cf[f'C{cf_row}'] = cash_flow_data['operating_activities']['net_cash']
        cf[f'C{cf_row}'].style = number_style
        cf_row += 2

        # B. Cash Flow from Investing Activities
        cf[f'A{cf_row}'] = "B. CASH FLOW FROM INVESTING ACTIVITIES"
        cf[f'A{cf_row}'].style = header_style
        cf_row += 1

        # Investing Activities Items
        investing_items = [
            ('Purchase of Fixed Assets', -cash_flow_data['investing_activities']['fixed_assets_purchase']),
            ('Sale of Fixed Assets', cash_flow_data['investing_activities']['fixed_assets_sale']),
            ('Purchase of Investments', -cash_flow_data['investing_activities']['investments_made']),
            ('Sale of Investments', cash_flow_data['investing_activities']['investments_sale'])
        ]

        for item in investing_items:
            if item[1] != 0:  # Only show non-zero items
                cf[f'A{cf_row}'] = item[0]
                cf[f'C{cf_row}'] = item[1]
                cf[f'C{cf_row}'].style = number_style
                cf_row += 1

        # Investing Activities Total
        cf[f'A{cf_row}'] = "Net Cash from Investing Activities (B)"
        cf[f'C{cf_row}'] = cash_flow_data['investing_activities']['net_cash']
        cf[f'C{cf_row}'].style = number_style
        cf_row += 2

        # C. Cash Flow from Financing Activities
        cf[f'A{cf_row}'] = "C. CASH FLOW FROM FINANCING ACTIVITIES"
        cf[f'A{cf_row}'].style = header_style
        cf_row += 1

        # Financing Activities Items
        financing_items = [
            ('Proceeds from Share Capital', cash_flow_data['financing_activities']['share_capital_proceeds']),
            ('Proceeds from Long-Term Borrowings', cash_flow_data['financing_activities']['long_term_borrowings']),
            ('Repayment of Long-Term Borrowings', cash_flow_data['financing_activities']['long_term_borrowings_repayment']),
            ('Net Changes in Short-Term Borrowings', cash_flow_data['financing_activities']['short_term_borrowings_net']),
            ('Interest Paid', cash_flow_data['financing_activities']['interest_paid']),
            ('Dividend Paid', cash_flow_data['financing_activities']['dividend_paid'])
        ]

        for item in financing_items:
            if item[1] != 0:  # Only show non-zero items
                cf[f'A{cf_row}'] = item[0]
                cf[f'C{cf_row}'] = item[1]
                cf[f'C{cf_row}'].style = number_style
                cf_row += 1

        # Financing Activities Total
        cf[f'A{cf_row}'] = "Net Cash from Financing Activities (C)"
        cf[f'C{cf_row}'] = cash_flow_data['financing_activities']['net_cash']
        cf[f'C{cf_row}'].style = number_style
        cf_row += 2

        # Net Change in Cash and Overall Position
        cf[f'A{cf_row}'] = "Net Increase/(Decrease) in Cash and Cash Equivalents (A+B+C)"
        cf[f'C{cf_row}'] = cash_flow_data['net_change']  # Changed from cash_position.net_increase
        cf[f'C{cf_row}'].style = number_style
        cf_row += 1

        # Opening and Closing Balance
        cf[f'A{cf_row}'] = "Cash and Cash Equivalents at the Beginning of the Year"
        cf[f'C{cf_row}'] = cash_flow_data['cash_beginning']  # Changed from cash_position.beginning_balance
        cf[f'C{cf_row}'].style = number_style
        cf_row += 1

        cf[f'A{cf_row}'] = "Cash and Cash Equivalents at the End of the Year"
        cf[f'C{cf_row}'] = cash_flow_data['cash_ending']  # Changed from cash_position.ending_balance
        cf[f'C{cf_row}'].style = number_style
        cf_row += 2

        # Add signature section similar to Balance Sheet and P&L
# Add signature section to Cash Flow Statement
        cf_row += 3  # Add some space before signatures

        # Header for signatures
        cf[f'A{cf_row}'] = "As per our report of even date attached"
        cf[f'A{cf_row}'].style = header_style
        cf.merge_cells(f'A{cf_row}:D{cf_row}')
        cf_row += 2

        # Store starting row for signature blocks
        signature_start_row = cf_row

        # Left Side - Auditor Details
        auditor_col = 'A'
        cf[f'{auditor_col}{cf_row}'] = f"For {client.get('audit_firm', {}).get('name', 'N/A')}"
        cf[f'{auditor_col}{cf_row}'].font = Font(bold=True)
        cf_row += 1

        cf[f'{auditor_col}{cf_row}'] = "Chartered Accountants"
        cf_row += 1

        cf[f'{auditor_col}{cf_row}'] = f"Firm Registration No: {client.get('audit_firm', {}).get('frn', 'N/A')}"
        cf_row += 2

        # Partner Details
        cf[f'{auditor_col}{cf_row}'] = client.get('audit_firm', {}).get('auditor_name', 'N/A')
        cf[f'{auditor_col}{cf_row}'].font = Font(bold=True)
        cf_row += 1

        cf[f'{auditor_col}{cf_row}'] = "Partner"
        cf_row += 1

        cf[f'{auditor_col}{cf_row}'] = f"Membership No: {client.get('audit_firm', {}).get('membership_no', 'N/A')}"
        cf_row += 1

        cf[f'{auditor_col}{cf_row}'] = f"UDIN: {client.get('audit_firm', {}).get('udin', 'N/A')}"
        cf_row += 1

        # Right Side - Client Details
        client_col = 'D'
        right_start = signature_start_row

        cf[f'{client_col}{right_start}'] = f"For {client.get('legal_name', 'N/A')}"
        cf[f'{client_col}{right_start}'].font = Font(bold=True)

        cf[f'{client_col}{right_start + 4}'] = client.get('signatory', {}).get('name', 'N/A')
        cf[f'{client_col}{right_start + 4}'].font = Font(bold=True)

        cf[f'{client_col}{right_start + 5}'] = client.get('signatory', {}).get('designation', 'N/A')

        cf[f'{client_col}{right_start + 6}'] = f"PAN: {client.get('signatory', {}).get('pan', 'N/A')}"

        # Place and Date
        cf_row += 2
        # Left side place and date
        cf[f'{auditor_col}{cf_row}'] = f"Place: {client.get('audit_firm', {}).get('place', 'N/A')}"
        cf_row += 1
        cf[f'{auditor_col}{cf_row}'] = f"Date: {format_date(client.get('audit_firm', {}).get('date'))}"

        # Right side place and date
        cf[f'{client_col}{cf_row-1}'] = f"Place: {client.get('signatory', {}).get('place', 'N/A')}"
        cf[f'{client_col}{cf_row}'] = f"Date: {format_date(client.get('signatory', {}).get('date'))}"        
        # Set column widths and print settings
        cf.column_dimensions['A'].width = 45
        cf.column_dimensions['B'].width = 8
        cf.column_dimensions['C'].width = 15
        cf.column_dimensions['D'].width = 15

        # Set print area and other formatting
        cf.print_title_rows = '1:5'
        cf.print_area = f'A1:D{cf_row}'
        cf.page_setup.orientation = cf.ORIENTATION_PORTRAIT
        cf.page_setup.paperSize = cf.PAPERSIZE_A4
        cf.page_setup.fitToPage = True
        cf.page_setup.fitToHeight = 1
        cf.page_setup.fitToWidth = 1
        cf.page_setup.horizontalCentered = True
        cf.page_setup.verticalCentered = True
        cf.print_options.gridLines = True

                # First add the setup_worksheet function before creating any sheets
        def setup_worksheet(ws, title):
            """Setup common worksheet formatting and headers"""
            ws.page_setup.orientation = ws.ORIENTATION_PORTRAIT
            ws.page_setup.paperSize = ws.PAPERSIZE_A4
            ws.page_setup.fitToPage = True
            ws.page_setup.fitToHeight = 1
            ws.page_setup.fitToWidth = 1
            
            # Set margins
            ws.page_margins.left = 0.5
            ws.page_margins.right = 0.5
            ws.page_margins.top = 0.5
            ws.page_margins.bottom = 0.5
            ws.page_margins.header = 0.3
            ws.page_margins.footer = 0.3

            # Set column widths
            ws.column_dimensions['A'].width = 45  # Particulars
            ws.column_dimensions['B'].width = 8   # Note
            ws.column_dimensions['C'].width = 15  # Current year
            ws.column_dimensions['D'].width = 15  # Previous year

            # Add headers
            ws['A1'] = f"{client['legal_name']} - {title}"
            ws['A1'].style = title_style
            ws.merge_cells('A1:D1')
            
            ws['A2'] = f"As at {balance_sheet_date.strftime('%B %d, %Y')}"
            ws['A2'].style = subheader_style
            ws.merge_cells('A2:D2')
            
            ws['A3'] = f"All amounts in {client['settings']['rounding_amount']}"
            ws.merge_cells('A3:D3')
            
            return 5  # Return starting row for content


        # Create Balance Sheet Notes
        # Generate balance sheet notes first
        notes_data = generate_balance_sheet_notes(trial_balance['data'], balance_sheet)

        # Create Balance Sheet Notes sheet
        notes = wb.create_sheet("Balance Sheet Notes")
        notes_row = setup_worksheet(notes, "Notes to Balance Sheet")

        # Headers
        headers = ["Particulars", "Note", f"As at {balance_sheet_date.strftime('%B %d, %Y')}", 
                f"As at {previous_year_date.strftime('%B %d, %Y')}"]
        for col, header in enumerate(headers, 1):
            notes.cell(row=notes_row, column=col, value=header).style = header_style
        notes_row += 2

        # Note 3: Shareholders' Funds
        notes[f'A{notes_row}'] = "Note 3: Shareholders' Funds"
        notes[f'A{notes_row}'].style = header_style
        notes.merge_cells(f'A{notes_row}:D{notes_row}')
        notes_row += 1

        # 3.1 Share Capital
        notes[f'A{notes_row}'] = "3.1 Share Capital"
        notes[f'A{notes_row}'].style = subheader_style
        notes_row += 1

        entries = balance_sheet['liabilities']['equity']['equity_share_capital']['entries']
        for entry in entries:
            notes[f'A{notes_row}'] = entry.get('account', '')
            notes[f'C{notes_row}'] = entry.get('amount', 0)
            notes[f'C{notes_row}'].style = number_style
            notes_row += 1

        # Share Capital Total
        notes[f'A{notes_row}'] = "Total Share Capital"
        notes[f'A{notes_row}'].font = Font(bold=True)
        notes[f'C{notes_row}'] = balance_sheet['liabilities']['equity']['equity_share_capital']['total']
        notes[f'C{notes_row}'].style = number_style
        notes_row += 2

        # 3.2 Reserves and Surplus
        notes[f'A{notes_row}'] = "3.2 Reserves and Surplus"
        notes[f'A{notes_row}'].style = subheader_style
        notes_row += 1

        # Process each reserve type
        reserve_types = [
            ('General Reserve', 'general_reserve'),
            ('Capital Reserve', 'capital_reserve'),
            ('Securities Premium', 'securities_premium'),
            ('Retained Earnings', 'retained_earnings')
        ]

        for title, key in reserve_types:
            notes[f'A{notes_row}'] = title
            notes[f'A{notes_row}'].font = Font(bold=True)
            notes_row += 1
            
            entries = balance_sheet['liabilities']['equity'][key]['entries']
            for entry in entries:
                notes[f'A{notes_row}'] = entry.get('account', '')
                notes[f'C{notes_row}'] = entry.get('amount', 0)
                notes[f'C{notes_row}'].style = number_style
                notes_row += 1
            
            if key == 'retained_earnings':
                # Add current year profit
                notes[f'A{notes_row}'] = "Add: Profit for the year"
                notes[f'C{notes_row}'] = balance_sheet['liabilities']['equity']['retained_earnings']['current_year_profit']
                notes[f'C{notes_row}'].style = number_style
                notes_row += 1
            
            notes[f'A{notes_row}'] = f"Total {title}"
            notes[f'C{notes_row}'] = balance_sheet['liabilities']['equity'][key]['total']
            notes[f'C{notes_row}'].style = number_style
            notes_row += 2

        # Total Reserves and Surplus
        total_reserves = sum(balance_sheet['liabilities']['equity'][key]['total'] for _, key in reserve_types)
        notes[f'A{notes_row}'] = "Total Reserves and Surplus"
        notes[f'A{notes_row}'].font = Font(bold=True)
        notes[f'C{notes_row}'] = total_reserves
        notes[f'C{notes_row}'].style = number_style
        notes_row += 2

        # Note 4: Long Term Borrowings
        notes[f'A{notes_row}'] = "Note 4: Long Term Borrowings"
        notes[f'A{notes_row}'].style = header_style
        notes.merge_cells(f'A{notes_row}:D{notes_row}')
        notes_row += 1

        for loan_type in ['secured_loans', 'unsecured_loans']:
            entries = balance_sheet['liabilities']['long_term_liabilities'][loan_type]['entries']
            title = "Secured Loans" if loan_type == 'secured_loans' else "Unsecured Loans"
            
            notes[f'A{notes_row}'] = title
            notes[f'A{notes_row}'].font = Font(bold=True)
            notes_row += 1
            
            for entry in entries:
                notes[f'A{notes_row}'] = entry.get('account', '')
                notes[f'C{notes_row}'] = entry.get('amount', 0)
                notes[f'C{notes_row}'].style = number_style
                notes_row += 1
            
            notes[f'A{notes_row}'] = f"Total {title}"
            notes[f'C{notes_row}'] = balance_sheet['liabilities']['long_term_liabilities'][loan_type]['total']
            notes[f'C{notes_row}'].style = number_style
            notes_row += 2

        # Note 5: Trade Payables
        notes[f'A{notes_row}'] = "Note 5: Trade Payables"
        notes[f'A{notes_row}'].style = header_style
        notes.merge_cells(f'A{notes_row}:D{notes_row}')
        notes_row += 1

        for payable_type in ['msme_payables', 'other_payables']:
            entries = balance_sheet['liabilities']['current_liabilities'][payable_type]['entries']
            title = "MSME Payables" if payable_type == 'msme_payables' else "Other Trade Payables"
            
            notes[f'A{notes_row}'] = title
            notes[f'A{notes_row}'].font = Font(bold=True)
            notes_row += 1
            
            for entry in entries:
                notes[f'A{notes_row}'] = entry.get('account', '')
                notes[f'C{notes_row}'] = entry.get('amount', 0)
                notes[f'C{notes_row}'].style = number_style
                notes_row += 1
            
            notes[f'A{notes_row}'] = f"Total {title}"
            notes[f'C{notes_row}'] = balance_sheet['liabilities']['current_liabilities'][payable_type]['total']
            notes[f'C{notes_row}'].style = number_style
            notes_row += 2

        # Note 6: Other Current Liabilities
        notes[f'A{notes_row}'] = "Note 6: Other Current Liabilities"
        notes[f'A{notes_row}'].style = header_style
        notes.merge_cells(f'A{notes_row}:D{notes_row}')
        notes_row += 1

        current_liab_types = {
            'statutory_dues': 'Statutory Dues',
            'short_term_borrowings': 'Short Term Borrowings',
            'other_current_liabilities': 'Other Current Liabilities'
        }

        for key, title in current_liab_types.items():
            entries = balance_sheet['liabilities']['current_liabilities'][key]['entries']
            
            notes[f'A{notes_row}'] = title
            notes[f'A{notes_row}'].font = Font(bold=True)
            notes_row += 1
            
            for entry in entries:
                notes[f'A{notes_row}'] = entry.get('account', '')
                notes[f'C{notes_row}'] = entry.get('amount', 0)
                notes[f'C{notes_row}'].style = number_style
                notes_row += 1
            
            notes[f'A{notes_row}'] = f"Total {title}"
            notes[f'C{notes_row}'] = balance_sheet['liabilities']['current_liabilities'][key]['total']
            notes[f'C{notes_row}'].style = number_style
            notes_row += 2

        # Continue with remaining notes...
        # Note 7: Fixed Assets
        notes[f'A{notes_row}'] = "Note 7: Fixed Assets"
        notes[f'A{notes_row}'].style = header_style
        notes.merge_cells(f'A{notes_row}:D{notes_row}')
        notes_row += 1

        for asset_type in ['tangible_assets', 'intangible_assets']:
            entries = balance_sheet['assets']['fixed_assets'][asset_type]['entries']
            title = "Tangible Assets" if asset_type == 'tangible_assets' else "Intangible Assets"
            
            notes[f'A{notes_row}'] = title
            notes[f'A{notes_row}'].font = Font(bold=True)
            notes_row += 1
            
            for entry in entries:
                notes[f'A{notes_row}'] = entry.get('account', '')
                notes[f'C{notes_row}'] = entry.get('amount', 0)
                notes[f'C{notes_row}'].style = number_style
                notes_row += 1
            
            notes[f'A{notes_row}'] = f"Total {title}"
            notes[f'C{notes_row}'] = balance_sheet['assets']['fixed_assets'][asset_type]['total']
            notes[f'C{notes_row}'].style = number_style
            notes_row += 2

            # Note 8: Investments
            notes[f'A{notes_row}'] = "Note 8: Investments"
            notes[f'A{notes_row}'].style = header_style
            notes.merge_cells(f'A{notes_row}:D{notes_row}')
            notes_row += 1

            # Long Term Investments
            notes[f'A{notes_row}'] = "Long Term Investments"
            notes[f'A{notes_row}'].font = Font(bold=True)
            notes_row += 1

            entries = balance_sheet['assets']['investments']['long_term_investments']['entries']
            for entry in entries:
                notes[f'A{notes_row}'] = entry.get('account', '')
                notes[f'C{notes_row}'] = entry.get('amount', 0)
                notes[f'C{notes_row}'].style = number_style
                notes_row += 1

            notes[f'A{notes_row}'] = "Total Long Term Investments"
            notes[f'C{notes_row}'] = balance_sheet['assets']['investments']['long_term_investments']['total']
            notes[f'C{notes_row}'].style = number_style
            notes_row += 2

            # Short Term Investments
            notes[f'A{notes_row}'] = "Short Term Investments"
            notes[f'A{notes_row}'].font = Font(bold=True)
            notes_row += 1

            entries = balance_sheet['assets']['investments']['short_term_investments']['entries']
            for entry in entries:
                notes[f'A{notes_row}'] = entry.get('account', '')
                notes[f'C{notes_row}'] = entry.get('amount', 0)
                notes[f'C{notes_row}'].style = number_style
                notes_row += 1

            notes[f'A{notes_row}'] = "Total Short Term Investments"
            notes[f'C{notes_row}'] = balance_sheet['assets']['investments']['short_term_investments']['total']
            notes[f'C{notes_row}'].style = number_style
            notes_row += 2

            # Total Investments
            notes[f'A{notes_row}'] = "Total Investments"
            notes[f'A{notes_row}'].font = Font(bold=True)
            notes[f'C{notes_row}'] = balance_sheet['assets']['investments']['total']
            notes[f'C{notes_row}'].style = number_style
            notes_row += 2

            # Note 9: Inventories
            notes[f'A{notes_row}'] = "Note 9: Inventories"
            notes[f'A{notes_row}'].style = header_style
            notes.merge_cells(f'A{notes_row}:D{notes_row}')
            notes_row += 1

            entries = balance_sheet['assets']['current_assets']['inventories']['entries']
            for entry in entries:
                notes[f'A{notes_row}'] = entry.get('account', '')
                notes[f'C{notes_row}'] = entry.get('amount', 0)
                notes[f'C{notes_row}'].style = number_style
                notes_row += 1

            notes[f'A{notes_row}'] = "Total Inventories"
            notes[f'A{notes_row}'].font = Font(bold=True)
            notes[f'C{notes_row}'] = balance_sheet['assets']['current_assets']['inventories']['total']
            notes[f'C{notes_row}'].style = number_style
            notes_row += 2

            # Note 10: Trade Receivables
            notes[f'A{notes_row}'] = "Note 10: Trade Receivables"
            notes[f'A{notes_row}'].style = header_style
            notes.merge_cells(f'A{notes_row}:D{notes_row}')
            notes_row += 1

            entries = balance_sheet['assets']['current_assets']['trade_receivables']['entries']
            for entry in entries:
                notes[f'A{notes_row}'] = entry.get('account', '')
                notes[f'C{notes_row}'] = entry.get('amount', 0)
                notes[f'C{notes_row}'].style = number_style
                notes_row += 1

            notes[f'A{notes_row}'] = "Total Trade Receivables"
            notes[f'A{notes_row}'].font = Font(bold=True)
            notes[f'C{notes_row}'] = balance_sheet['assets']['current_assets']['trade_receivables']['total']
            notes[f'C{notes_row}'].style = number_style
            notes_row += 2

            # Note 11: Cash and Cash Equivalents
            notes[f'A{notes_row}'] = "Note 11: Cash and Cash Equivalents"
            notes[f'A{notes_row}'].style = header_style
            notes.merge_cells(f'A{notes_row}:D{notes_row}')
            notes_row += 1

            entries = balance_sheet['assets']['current_assets']['cash_and_equivalents']['entries']
            for entry in entries:
                notes[f'A{notes_row}'] = entry.get('account', '')
                notes[f'C{notes_row}'] = entry.get('amount', 0)
                notes[f'C{notes_row}'].style = number_style
                notes_row += 1

            notes[f'A{notes_row}'] = "Total Cash and Cash Equivalents"
            notes[f'A{notes_row}'].font = Font(bold=True)
            notes[f'C{notes_row}'] = balance_sheet['assets']['current_assets']['cash_and_equivalents']['total']
            notes[f'C{notes_row}'].style = number_style
            notes_row += 2

            # Note 12: Other Current Assets
            notes[f'A{notes_row}'] = "Note 12: Other Current Assets"
            notes[f'A{notes_row}'].style = header_style
            notes.merge_cells(f'A{notes_row}:D{notes_row}')
            notes_row += 1

            # Loans and Advances
            notes[f'A{notes_row}'] = "Loans and Advances"
            notes[f'A{notes_row}'].font = Font(bold=True)
            notes_row += 1

            entries = balance_sheet['assets']['current_assets']['loans_and_advances']['entries']
            for entry in entries:
                notes[f'A{notes_row}'] = entry.get('account', '')
                notes[f'C{notes_row}'] = entry.get('amount', 0)
                notes[f'C{notes_row}'].style = number_style
                notes_row += 1

            notes[f'A{notes_row}'] = "Total Loans and Advances"
            notes[f'C{notes_row}'] = balance_sheet['assets']['current_assets']['loans_and_advances']['total']
            notes[f'C{notes_row}'].style = number_style
            notes_row += 2

            # Other Current Assets
            notes[f'A{notes_row}'] = "Other Current Assets"
            notes[f'A{notes_row}'].font = Font(bold=True)
            notes_row += 1

            entries = balance_sheet['assets']['current_assets']['other_current_assets']['entries']
            for entry in entries:
                notes[f'A{notes_row}'] = entry.get('account', '')
                notes[f'C{notes_row}'] = entry.get('amount', 0)
                notes[f'C{notes_row}'].style = number_style
                notes_row += 1

            notes[f'A{notes_row}'] = "Total Other Current Assets"
            notes[f'C{notes_row}'] = balance_sheet['assets']['current_assets']['other_current_assets']['total']
            notes[f'C{notes_row}'].style = number_style
            notes_row += 1

            # Total for Note 12
            notes[f'A{notes_row}'] = "Total Other Current Assets (Note 12)"
            notes[f'A{notes_row}'].font = Font(bold=True)
            total_other_current = (
                balance_sheet['assets']['current_assets']['loans_and_advances']['total'] +
                balance_sheet['assets']['current_assets']['other_current_assets']['total']
            )
            notes[f'C{notes_row}'] = total_other_current
            notes[f'C{notes_row}'].style = number_style
            notes[f'C{notes_row}'].font = Font(bold=True)
            notes_row += 2

        # Add signature section
        notes_row += 2
        notes[f'A{notes_row}'] = "As per our report of even date attached"
        notes[f'A{notes_row}'].style = header_style
        notes.merge_cells(f'A{notes_row}:D{notes_row}')
        notes_row += 2

        # Left Side - Auditor Details
        auditor_col = 'A'
        notes[f'{auditor_col}{notes_row}'] = f"For {client.get('audit_firm', {}).get('name', 'N/A')}"
        notes[f'{auditor_col}{notes_row}'].font = Font(bold=True)
        notes_row += 1

        notes[f'{auditor_col}{notes_row}'] = "Chartered Accountants"
        notes_row += 1

        notes[f'{auditor_col}{notes_row}'] = f"Firm Registration No: {client.get('audit_firm', {}).get('frn', 'N/A')}"
        notes_row += 2

        notes[f'{auditor_col}{notes_row}'] = client.get('audit_firm', {}).get('auditor_name', 'N/A')
        notes[f'{auditor_col}{notes_row}'].font = Font(bold=True)
        notes_row += 1

        notes[f'{auditor_col}{notes_row}'] = "Partner"
        notes_row += 1

        notes[f'{auditor_col}{notes_row}'] = f"Membership No: {client.get('audit_firm', {}).get('membership_no', 'N/A')}"
        notes_row += 1

        # Right Side - Client Details
        client_col = 'D'
        signature_start = notes_row - 4
        notes[f'{client_col}{signature_start}'] = f"For {client.get('legal_name', 'N/A')}"
        notes[f'{client_col}{signature_start}'].font = Font(bold=True)

        notes[f'{client_col}{signature_start + 2}'] = client.get('signatory', {}).get('name', 'N/A')
        notes[f'{client_col}{signature_start + 2}'].font = Font(bold=True)
        notes[f'{client_col}{signature_start + 3}'] = client.get('signatory', {}).get('designation', 'N/A')

        # Place and Date
        notes_row += 2
        notes[f'{auditor_col}{notes_row}'] = f"Place: {client.get('audit_firm', {}).get('place', 'N/A')}"
        notes[f'{client_col}{notes_row}'] = f"Place: {client.get('signatory', {}).get('place', 'N/A')}"
        notes_row += 1
        notes[f'{auditor_col}{notes_row}'] = f"Date: {format_date(client.get('audit_firm', {}).get('date'))}"
        notes[f'{client_col}{notes_row}'] = f"Date: {format_date(client.get('signatory', {}).get('date'))}"

        # Set print area and other formatting
        notes.print_title_rows = '1:5'
        notes.print_area = f'A1:D{notes_row}'
        notes.page_setup.orientation = notes.ORIENTATION_PORTRAIT
        notes.page_setup.paperSize = notes.PAPERSIZE_A4
        notes.page_setup.fitToPage = True
        notes.page_setup.fitToHeight = 1
        notes.page_setup.fitToWidth = 1
        notes.page_setup.horizontalCentered = True
        notes.page_setup.verticalCentered = True
        notes.print_options.gridLines = True

        # Set column widths
        notes.column_dimensions['A'].width = 45
        notes.column_dimensions['B'].width = 8
        notes.column_dimensions['C'].width = 15
        notes.column_dimensions['D'].width = 15

        # After your Balance Sheet Notes section, add this code:

        # Create Profit and Loss Notes sheet
        pl_notes = wb.create_sheet("Profit and Loss Notes")

        # Setup worksheet with common settings
        pl_notes_row = setup_worksheet(pl_notes, "Notes to Profit and Loss Statement")

        # Headers
        headers = ["Particulars", "Note", f"Year ended {balance_sheet_date.strftime('%B %d, %Y')}", 
                f"Year ended {previous_year_date.strftime('%B %d, %Y')}"]
        for col, header in enumerate(headers, 1):
            pl_notes.cell(row=pl_notes_row, column=col, value=header).style = header_style
        pl_notes_row += 2

        # Note 12: Revenue from Operations
        pl_notes[f'A{pl_notes_row}'] = "Note 12: Revenue from Operations"
        pl_notes[f'A{pl_notes_row}'].style = header_style
        pl_notes.merge_cells(f'A{pl_notes_row}:D{pl_notes_row}')
        pl_notes_row += 1

        for item in profit_loss_notes['note_12']['entries']:
            pl_notes[f'A{pl_notes_row}'] = item.get('account', '')
            pl_notes[f'C{pl_notes_row}'] = item.get('amount', 0)
            pl_notes[f'C{pl_notes_row}'].style = number_style
            pl_notes_row += 1

        pl_notes[f'A{pl_notes_row}'] = "Total Revenue from Operations"
        pl_notes[f'A{pl_notes_row}'].font = Font(bold=True)
        pl_notes[f'C{pl_notes_row}'] = profit_loss_notes['note_12']['total']
        pl_notes[f'C{pl_notes_row}'].style = number_style
        pl_notes[f'C{pl_notes_row}'].font = Font(bold=True)
        pl_notes_row += 2

        # Note 13: Other Income
        pl_notes[f'A{pl_notes_row}'] = "Note 13: Other Income"
        pl_notes[f'A{pl_notes_row}'].style = header_style
        pl_notes.merge_cells(f'A{pl_notes_row}:D{pl_notes_row}')
        pl_notes_row += 1

        for item in profit_loss_notes['note_13']['entries']:
            pl_notes[f'A{pl_notes_row}'] = item.get('account', '')
            pl_notes[f'C{pl_notes_row}'] = item.get('amount', 0)
            pl_notes[f'C{pl_notes_row}'].style = number_style
            pl_notes_row += 1

        pl_notes[f'A{pl_notes_row}'] = "Total Other Income"
        pl_notes[f'A{pl_notes_row}'].font = Font(bold=True)
        pl_notes[f'C{pl_notes_row}'] = profit_loss_notes['note_13']['total']
        pl_notes[f'C{pl_notes_row}'].style = number_style
        pl_notes[f'C{pl_notes_row}'].font = Font(bold=True)
        pl_notes_row += 2

        # Note 14: Cost of Materials
        pl_notes[f'A{pl_notes_row}'] = "Note 14: Cost of Materials Consumed"
        pl_notes[f'A{pl_notes_row}'].style = header_style
        pl_notes.merge_cells(f'A{pl_notes_row}:D{pl_notes_row}')
        pl_notes_row += 1

        # Opening Stock
        pl_notes[f'A{pl_notes_row}'] = "Opening Stock"
        opening_stock = sum(item.get('amount', 0) for item in profit_loss_notes['note_14']['entries'] 
                        if 'Opening' in item.get('account', ''))
        pl_notes[f'C{pl_notes_row}'] = opening_stock
        pl_notes[f'C{pl_notes_row}'].style = number_style
        pl_notes_row += 1

        # Add Purchases
        pl_notes[f'A{pl_notes_row}'] = "Add: Purchases"
        purchases = sum(item.get('amount', 0) for item in profit_loss_notes['note_14']['entries'] 
                    if 'Purchase' in item.get('account', ''))
        pl_notes[f'C{pl_notes_row}'] = purchases
        pl_notes[f'C{pl_notes_row}'].style = number_style
        pl_notes_row += 1

        # Less Closing Stock
        pl_notes[f'A{pl_notes_row}'] = "Less: Closing Stock"
        closing_stock = sum(item.get('amount', 0) for item in profit_loss_notes['note_14']['entries'] 
                        if 'Closing' in item.get('account', ''))
        pl_notes[f'C{pl_notes_row}'] = -closing_stock
        pl_notes[f'C{pl_notes_row}'].style = number_style
        pl_notes_row += 1

        # Total Cost of Materials
        pl_notes[f'A{pl_notes_row}'] = "Total Cost of Materials Consumed"
        pl_notes[f'A{pl_notes_row}'].font = Font(bold=True)
        pl_notes[f'C{pl_notes_row}'] = profit_loss_notes['note_14']['total']
        pl_notes[f'C{pl_notes_row}'].style = number_style
        pl_notes[f'C{pl_notes_row}'].font = Font(bold=True)
        pl_notes_row += 2

        # Note 15: Employee Benefits Expense
        pl_notes[f'A{pl_notes_row}'] = "Note 15: Employee Benefits Expense"
        pl_notes[f'A{pl_notes_row}'].style = header_style
        pl_notes.merge_cells(f'A{pl_notes_row}:D{pl_notes_row}')
        pl_notes_row += 1

        # Process each subcategory of Employee Benefits
        for category, data in profit_loss_notes['note_15']['sub_categories'].items():
            # Category header
            pl_notes[f'A{pl_notes_row}'] = data['title']
            pl_notes[f'A{pl_notes_row}'].font = Font(bold=True)
            pl_notes_row += 1

            # Category entries
            for item in data['entries']:
                pl_notes[f'A{pl_notes_row}'] = item.get('account', '')
                pl_notes[f'C{pl_notes_row}'] = item.get('amount', 0)
                pl_notes[f'C{pl_notes_row}'].style = number_style
                pl_notes_row += 1

            # Subtotal for category
            pl_notes[f'A{pl_notes_row}'] = f"Sub-total: {data['title']}"
            pl_notes[f'A{pl_notes_row}'].font = Font(italic=True)
            pl_notes[f'C{pl_notes_row}'] = data['total']
            pl_notes[f'C{pl_notes_row}'].style = number_style
            pl_notes_row += 1

        # Total Employee Benefits
        pl_notes[f'A{pl_notes_row}'] = "Total Employee Benefits Expense"
        pl_notes[f'A{pl_notes_row}'].font = Font(bold=True)
        pl_notes[f'C{pl_notes_row}'] = profit_loss_notes['note_15']['total']
        pl_notes[f'C{pl_notes_row}'].style = number_style
        pl_notes_row += 2

        # Note 16: Finance Costs
        pl_notes[f'A{pl_notes_row}'] = "Note 16: Finance Costs"
        pl_notes[f'A{pl_notes_row}'].style = header_style
        pl_notes.merge_cells(f'A{pl_notes_row}:D{pl_notes_row}')
        pl_notes_row += 1

        # Process each subcategory of Finance Costs
        for category, data in profit_loss_notes['note_16']['sub_categories'].items():
            # Category header
            pl_notes[f'A{pl_notes_row}'] = data['title']
            pl_notes[f'A{pl_notes_row}'].font = Font(bold=True)
            pl_notes_row += 1

            # Category entries
            for item in data['entries']:
                pl_notes[f'A{pl_notes_row}'] = item.get('account', '')
                pl_notes[f'C{pl_notes_row}'] = item.get('amount', 0)
                pl_notes[f'C{pl_notes_row}'].style = number_style
                pl_notes_row += 1

            # Subtotal for category
            pl_notes[f'A{pl_notes_row}'] = f"Sub-total: {data['title']}"
            pl_notes[f'A{pl_notes_row}'].font = Font(italic=True)
            pl_notes[f'C{pl_notes_row}'] = data['total']
            pl_notes[f'C{pl_notes_row}'].style = number_style
            pl_notes_row += 1

        # Total Finance Costs
        pl_notes[f'A{pl_notes_row}'] = "Total Finance Costs"
        pl_notes[f'A{pl_notes_row}'].font = Font(bold=True)
        pl_notes[f'C{pl_notes_row}'] = profit_loss_notes['note_16']['total']
        pl_notes[f'C{pl_notes_row}'].style = number_style
        pl_notes_row += 2

        # Note 17: Other Expenses
        pl_notes[f'A{pl_notes_row}'] = "Note 17: Other Expenses"
        pl_notes[f'A{pl_notes_row}'].style = header_style
        pl_notes.merge_cells(f'A{pl_notes_row}:D{pl_notes_row}')
        pl_notes_row += 1

        for item in profit_loss_notes['note_17']['entries']:
            pl_notes[f'A{pl_notes_row}'] = item.get('account', '')
            pl_notes[f'C{pl_notes_row}'] = item.get('amount', 0)
            pl_notes[f'C{pl_notes_row}'].style = number_style
            pl_notes_row += 1

        # Other Expenses Total
        pl_notes[f'A{pl_notes_row}'] = "Total Other Expenses"
        pl_notes[f'A{pl_notes_row}'].font = Font(bold=True)
        pl_notes[f'C{pl_notes_row}'] = profit_loss_notes['note_17']['total']
        pl_notes[f'C{pl_notes_row}'].style = number_style
        pl_notes_row += 2


        # Add signature section
        pl_notes_row += 2
        pl_notes[f'A{pl_notes_row}'] = "As per our report of even date attached"
        pl_notes[f'A{pl_notes_row}'].style = header_style
        pl_notes.merge_cells(f'A{pl_notes_row}:D{pl_notes_row}')
        pl_notes_row += 2

        # Left Side - Auditor Details
        auditor_col = 'A'
        pl_notes[f'{auditor_col}{pl_notes_row}'] = f"For {client.get('audit_firm', {}).get('name', 'N/A')}"
        pl_notes[f'{auditor_col}{pl_notes_row}'].font = Font(bold=True)
        pl_notes_row += 1

        pl_notes[f'{auditor_col}{pl_notes_row}'] = "Chartered Accountants"
        pl_notes_row += 1

        pl_notes[f'{auditor_col}{pl_notes_row}'] = f"Firm Registration No: {client.get('audit_firm', {}).get('frn', 'N/A')}"
        pl_notes_row += 2

        pl_notes[f'{auditor_col}{pl_notes_row}'] = client.get('audit_firm', {}).get('auditor_name', 'N/A')
        pl_notes[f'{auditor_col}{pl_notes_row}'].font = Font(bold=True)
        pl_notes_row += 1

        pl_notes[f'{auditor_col}{pl_notes_row}'] = "Partner"
        pl_notes_row += 1

        pl_notes[f'{auditor_col}{pl_notes_row}'] = f"Membership No: {client.get('audit_firm', {}).get('membership_no', 'N/A')}"
        pl_notes_row += 1

        # Right Side - Client Details
        client_col = 'D'
        signature_start = pl_notes_row - 4
        pl_notes[f'{client_col}{signature_start}'] = f"For {client.get('legal_name', 'N/A')}"
        pl_notes[f'{client_col}{signature_start}'].font = Font(bold=True)

        pl_notes[f'{client_col}{signature_start + 2}'] = client.get('signatory', {}).get('name', 'N/A')
        pl_notes[f'{client_col}{signature_start + 2}'].font = Font(bold=True)
        pl_notes[f'{client_col}{signature_start + 3}'] = client.get('signatory', {}).get('designation', 'N/A')

        # Place and Date
        pl_notes_row += 2
        pl_notes[f'{auditor_col}{pl_notes_row}'] = f"Place: {client.get('audit_firm', {}).get('place', 'N/A')}"
        pl_notes[f'{client_col}{pl_notes_row}'] = f"Place: {client.get('signatory', {}).get('place', 'N/A')}"
        pl_notes_row += 1
        pl_notes[f'{auditor_col}{pl_notes_row}'] = f"Date: {format_date(client.get('audit_firm', {}).get('date'))}"
        pl_notes[f'{client_col}{pl_notes_row}'] = f"Date: {format_date(client.get('signatory', {}).get('date'))}"

        # Set print area and other formatting
        pl_notes.print_title_rows = '1:5'
        pl_notes.print_area = f'A1:D{pl_notes_row}'
        pl_notes.page_setup.orientation = pl_notes.ORIENTATION_PORTRAIT
        pl_notes.page_setup.paperSize = pl_notes.PAPERSIZE_A4
        pl_notes.page_setup.fitToPage = True
        pl_notes.page_setup.fitToHeight = 1
        pl_notes.page_setup.fitToWidth = 1
        pl_notes.page_setup.horizontalCentered = True
        pl_notes.page_setup.verticalCentered = True
        pl_notes.print_options.gridLines = True

        # Then continue with your existing code for file creation and response
        excel_file = BytesIO()
        wb.save(excel_file)
        excel_file.seek(0)

        

        
        return send_file(
            excel_file,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=f'Financial_Statements_{client["legal_name"]}_{client["fiscal_year"]}.xlsx'
        )
        
    except Exception as e:
        print(f"Error exporting financials: {str(e)}")
        traceback.print_exc()
        flash('Error exporting financial statements', 'error')
        return redirect(url_for('financials.view_client_details', client_id=client_id))
